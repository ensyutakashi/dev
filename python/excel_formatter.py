# --- obsidian_property ---
# scr名: 【自動】
# 概要: Excel表整形
# 処理grp: tool
# 処理順番: 0
# mermaid: -
# input: 無し
# output: 無し
# tags: ["tool", "excel"]
# aliases: ["excel_formatter.py"]
# created: 2026-03-19
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：Excel表整形
"""
font:源ノ角ゴシック Code JP R
background:white
列幅オートフィット
header:太字
先頭行の固定
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# 定数
DEFAULT_FONT_NAME = "源ノ角ゴシック Code JP R"
FALLBACK_FONT_NAME = "Yu Gothic"
DEFAULT_FONT_SIZE = 10
DEFAULT_CELL_COLOR = "FFFFFF"
MAX_COLUMN_WIDTH = 50
COLUMN_WIDTH_PADDING = 2


class ExcelFormatter:
    """Excelファイルを整形するクラス"""
    
    def __init__(self, font_name: Optional[str] = None):
        """フォーマッターを初期化
        
        Args:
            font_name: 使用するフォント名（Noneの場合はデフォルトを使用）
        """
        self.font_name = font_name or DEFAULT_FONT_NAME
        self.font = self._create_font()
        self.white_fill = PatternFill("solid", fgColor=DEFAULT_CELL_COLOR)
        self.header_font = self._create_font(bold=True)
    
    def _create_font(self, bold: bool = False) -> Font:
        """フォントオブジェクトを作成
        
        Args:
            bold: 太字の場合はTrue
            
        Returns:
            Font: フォントオブジェクト
        """
        try:
            return Font(name=self.font_name, size=DEFAULT_FONT_SIZE, bold=bold)
        except Exception:
            # フォントが見つからない場合はフォールバック
            return Font(name=FALLBACK_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=bold)
    
    def _format_sheet(self, sheet) -> None:
        """シートを整形
        
        Args:
            sheet: 整形対象のシート
        """
        sheet_name = sheet.title
        print(f"[INFO] シート「{sheet_name}」を整形中...")
        
        # セルの整形
        for row in sheet.iter_rows():
            for cell in row:
                # 背景を白に設定
                cell.fill = self.white_fill
                
                # フォント設定
                if cell.row == 1:
                    # ヘッダー行は太字
                    cell.font = self.header_font
                else:
                    # データ行は通常フォント
                    cell.font = self.font
        
        # 列幅のオートフィット
        self._auto_fit_columns(sheet)
        
        # グリッド線を非表示
        sheet.sheet_view.showGridLines = False
        
        # 先頭行を固定
        sheet.freeze_panes = "A2"
        
        print(f"[OK] シート「{sheet_name}」の整形完了")
    
    def _auto_fit_columns(self, sheet) -> None:
        """列幅を自動調整
        
        Args:
            sheet: 対象シート
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # 列内の最大文字長を取得
            for cell in column:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            # 列幅を設定（余裕を持たせ、最大幅を制限）
            adjusted_width = min(max_length + COLUMN_WIDTH_PADDING, MAX_COLUMN_WIDTH)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def format_excel(self, input_path: Path, output_path: Path, all_sheets: bool = False) -> None:
        """Excelファイルを整形
        
        Args:
            input_path: 入力Excelファイルパス
            output_path: 出力Excelファイルパス
            all_sheets: 全シートを整形する場合はTrue、アクティブシートのみの場合はFalse
        """
        # 入力ファイルを読み込み
        wb = load_workbook(input_path)
        
        # 整形対象のシートを決定
        if all_sheets:
            target_sheets = wb.worksheets
            print(f"[INFO] 全{len(target_sheets)}シートを整形します")
        else:
            target_sheets = [wb.active]
            print(f"[INFO] アクティブシートのみを整形します")
        
        # 各シートを整形
        for sheet in target_sheets:
            self._format_sheet(sheet)
        
        # 出力ファイルを保存
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)


def format_excel(input_path: Path, output_path: Path, all_sheets: bool = False) -> None:
    """Excelファイルを整形する（互換性のための関数）
    
    Args:
        input_path: 入力Excelファイルパス
        output_path: 出力Excelファイルパス
        all_sheets: 全シートを整形する場合はTrue、アクティブシートのみの場合はFalse
    """
    formatter = ExcelFormatter()
    formatter.format_excel(input_path, output_path, all_sheets)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Excelファイルを整形する")
    parser.add_argument("--input", required=True, help="入力Excelファイル")
    parser.add_argument("--output", required=True, help="出力Excelファイル")
    parser.add_argument("--all-sheets", action="store_true", 
                       help="全シートを整形する（デフォルトはアクティブシートのみ）")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    
    input_path = Path(args.input)
    output_path = Path(args.output)
    
    if not input_path.exists():
        print(f"[ERROR] 入力ファイルが見つかりません: {input_path}")
        return 1
    
    try:
        format_excel(input_path, output_path, all_sheets=args.all_sheets)
        sheet_info = "全シート" if args.all_sheets else "アクティブシート"
        print(f"[OK] Excelを{sheet_info}整形しました: {output_path}")
        return 0
    except Exception as exc:
        print(f"[ERROR] 整形に失敗しました: {exc}")
        return 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
