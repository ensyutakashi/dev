# --- obsidian_property ---
# scr名: 【自動】
# 概要: TDnet,PDF/XBRL,DB差分チェック
# 処理grp: TDnetダウンロード
# 処理順番: 6
# mermaid: "[[mermaid_TDnet適時開示情報ダウンロード]]"
# tags: ["tdnet", "download", "diff"]
# aliases: ["06_tdnet_folderfile_count.py"]
# created: 2026-02-17
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：？？？？
# --- 概要 ---


import os
import time
import re
import requests
import pandas as pd # type: ignore
import duckdb
from datetime import datetime, timedelta
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from concurrent.futures import ThreadPoolExecutor

# --- 設定項目 ---
BASE_PATH_PDF = r"\\LS720D7A9\TakashiBK\投資\TDNET\TDnet適時情報開示サービス\TDnet(決算短信)PDF"
BASE_PATH_XBRL = r"\\LS720D7A9\TakashiBK\投資\TDNET\TDnet適時情報開示サービス\TDnet(決算短信)XBRL"
DB_PATH = r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス\tdnet.duckdb"
DB_NAME = "tdnet.duckdb"  # DB名を指定する変数

TARGET_YEARS = ["2025", "2026"]
START_DATE = "2025-07-01"
END_DATE = datetime.now().strftime('%Y-%m-%d')  # 今日まで

# Excel出力先フォルダ設定（初期値：スクリプトと同じ場所）
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
# 環境変数から作業フォルダを取得（なければデフォルト）
WORKING_DIR = os.environ.get('TDNET_WORKING_DIR', os.path.join(OUTPUT_DIR, "TDnet_report_temp_validation_log_files"))
os.makedirs(WORKING_DIR, exist_ok=True)  # フォルダがなければ作成
OUTPUT_EXCEL = os.path.join(WORKING_DIR, f"tdnet_file_counts_summary_{datetime.now().strftime('%y%m%d_%H%M%S')}.xlsx")
FONT_NAME = "源ノ角ゴシック Code JP R"

# タブ名を指定する変数
TAB_FILE_COUNT = "ファイル数"
TAB_FILENAME_CHECK = "ファイル名"

TDNET_LIST_URL = "https://www.release.tdnet.info/inbs/I_list_001_"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
}
# ---------------

def get_date_from_filename(filename):
    try:
        date_str = filename[7:13]
        if date_str.isdigit() and len(date_str) == 6:
            return f"20{date_str[0:2]}-{date_str[2:4]}-{date_str[4:6]}"
    except: pass
    return None

def get_folder_files():
    """フォルダ内のファイル名を取得する"""
    folder_files = []
    
    # PDFファイル
    for year in TARGET_YEARS:
        year_dir = os.path.join(BASE_PATH_PDF, f"{year}年")
        if not os.path.exists(year_dir): continue
        for month in range(1, 13):
            month_str = f"{month:02}月"
            target_folder = os.path.join(year_dir, f"TDnet(決算短信)PDF{year}年{month_str}")
            if os.path.exists(target_folder) and os.path.isdir(target_folder):
                for file in os.listdir(target_folder):
                    if file.lower().endswith('.pdf'):
                        # 拡張子を削除してファイル名を登録
                        filename_without_ext = file[:-4]  # .pdfを削除
                        folder_files.append({"種類": "PDF", "ファイル名": filename_without_ext})
    
    # XBRLファイル
    for year in TARGET_YEARS:
        year_dir = os.path.join(BASE_PATH_XBRL, f"{year}年")
        if not os.path.exists(year_dir): continue
        for month in range(1, 13):
            month_str = f"{month:02}月"
            target_folder = os.path.join(year_dir, f"TDnet(決算短信)XBRL{year}年{month_str}")
            if os.path.exists(target_folder) and os.path.isdir(target_folder):
                for file in os.listdir(target_folder):
                    if file.lower().endswith('.zip'):
                        # 拡張子を削除してファイル名を登録
                        filename_without_ext = file[:-4]  # .zipを削除
                        folder_files.append({"種類": "XBRL", "ファイル名": filename_without_ext})
    
    return pd.DataFrame(folder_files)

def get_db_files():
    """DBからファイル名と関連情報を取得する"""
    try:
        con = duckdb.connect(DB_PATH)
        
        # PDFファイル
        df_pdf = con.execute("""
            SELECT "ファイル名(連番+公開日+時刻+(種別)+決算月+4Q+コード+会社名+表題)" as ファイル名, 
                   "時刻" as 公開日,
                   "連番" as 連番,
                   "pdfDL" as pdfDL,
                   "xbrlDL" as xbrlDL
            FROM disclosure_info
            WHERE CAST("時刻" AS DATE) BETWEEN '2025-01-01' AND '2026-12-31'
        """).df()
        df_pdf["種類"] = "PDF"
        
        # XBRLファイル（XBRL列が'XBRL'のもののみ）
        df_xbrl = con.execute("""
            SELECT "ファイル名(連番+公開日+時刻+(種別)+決算月+4Q+コード+会社名+表題)" as ファイル名, 
                   "時刻" as 公開日,
                   "連番" as 連番,
                   "pdfDL" as pdfDL,
                   "xbrlDL" as xbrlDL
            FROM disclosure_info
            WHERE CAST("時刻" AS DATE) BETWEEN '2025-01-01' AND '2026-12-31'
            AND "XBRL" = 'XBRL'
        """).df()
        df_xbrl["種類"] = "XBRL"
        
        con.close()
        
        # 結合
        df_db = pd.concat([df_pdf, df_xbrl], ignore_index=True)
        return df_db[["種類", "ファイル名", "公開日", "連番", "pdfDL", "xbrlDL"]]
        
    except Exception as e:
        print(f"   [DBエラー]: {e}")
        return pd.DataFrame(columns=["種類", "ファイル名", "公開日", "連番", "pdfDL", "xbrlDL"])

def compare_filenames():
    """DBとフォルダのファイル名を比較する"""
    print("ファイル名比較を開始します...")
    
    # フォルダとDBのファイル名を取得
    df_folder = get_folder_files()
    df_db = get_db_files()
    
    # 比較処理
    results = []
    
    # DBにしかないファイル
    db_only = pd.merge(df_db, df_folder, on=["種類", "ファイル名"], how="left", indicator=True)
    db_only = db_only[db_only["_merge"] == "left_only"]
    for _, row in db_only.iterrows():
        results.append({
            "種類": row["種類"], 
            "ファイル名": row["ファイル名"], 
            "公開日": row["公開日"],
            "連番": row["連番"],
            "pdfDL": row["pdfDL"],
            "xbrlDL": row["xbrlDL"],
            "判定": "フォルダになし"
        })
    
    # フォルダにしかないファイル
    folder_only = pd.merge(df_folder, df_db, on=["種類", "ファイル名"], how="left", indicator=True)
    folder_only = folder_only[folder_only["_merge"] == "left_only"]
    for _, row in folder_only.iterrows():
        results.append({
            "種類": row["種類"], 
            "ファイル名": row["ファイル名"], 
            "公開日": "",
            "連番": "",
            "pdfDL": "",
            "xbrlDL": "",
            "判定": "DBになし"
        })
    
    return pd.DataFrame(results)

def fetch_tdnet_count_single(d_str, valid_start_date):
    target_dt = datetime.strptime(d_str, '%Y-%m-%d').date()
    today_dt = datetime.now().date()
    if target_dt < valid_start_date or target_dt > today_dt:
        return d_str, 0
    date_param = d_str.replace("-", "")
    url = f"{TDNET_LIST_URL}{date_param}.html"
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        response.encoding = response.apparent_encoding
        html = response.text
        if "に開示された情報はありません。" in html:
            return d_str, 0
        match = re.search(r'全(\d+)件', html)
        if match:
            return d_str, int(match.group(1))
    except: pass
    return d_str, 0

def process_all_data():
    start_total_time = time.time()
    print(f"開始時刻: {datetime.now().strftime('%H:%M:%S')}")
    
    valid_start_date = datetime.now().date() - timedelta(days=30)

    try:
        con = duckdb.connect(database=DB_PATH, read_only=True)
        con.execute("SELECT 1")
        con.close()
    except Exception as e:
        msg = str(e)
        is_locked = (
            "already open" in msg.lower()
            or "使用中" in msg
            or "アクセスできません" in msg
            or "access" in msg.lower()
        )
        if is_locked:
            print(f"[警告] DBがロックされています。DBeaverなどで {DB_PATH} を閉じてから再実行してください。")
        else:
            print(f"[DBエラー] {e}")
        return

    # 1. フォルダ走査
    print("1/4: フォルダ走査を開始します...")
    step_start = time.time()
    folder_records = []
    targets = [("pdf_folder", BASE_PATH_PDF, ".pdf"), ("folder_XBRL", BASE_PATH_XBRL, ".zip")]
    
    for label, root_path, ext in targets:
        for year in TARGET_YEARS:
            year_dir = os.path.join(root_path, f"{year}年")
            if not os.path.exists(year_dir): continue
            for month in range(1, 13):
                month_str = f"{month:02}月"
                type_name = "PDF" if "pdf_folder" in label else "XBRL"
                target_folder = os.path.join(year_dir, f"TDnet(決算短信){type_name}{year}年{month_str}")
                if os.path.exists(target_folder) and os.path.isdir(target_folder):
                    for file in sorted(os.listdir(target_folder)):
                        if file.lower().endswith(ext.lower()):
                            d = get_date_from_filename(file)
                            if d: folder_records.append({"date": d, "type": label})
    
    df_folder_raw = pd.DataFrame(folder_records)
    df_folder_agg = df_folder_raw.groupby(['date', 'type']).size().unstack(fill_value=0).reset_index() if not df_folder_raw.empty else pd.DataFrame(columns=['date', 'pdf_folder', 'folder_XBRL'])
    print(f"   -> フォルダ走査完了 ({time.time() - step_start:.2f}秒)")

    # 2. DuckDB取得
    print("2/4: DuckDBからデータを取得します...")
    step_start = time.time()
    try:
        con = duckdb.connect(DB_PATH)
        df_db = con.execute(f"""
            SELECT strftime(CAST("時刻" AS DATE), '%Y-%m-%d') as date,
                   COUNT(*) as pdf_db,
                   COUNT(CASE WHEN "XBRL" = 'XBRL' THEN 1 END) as db_XBRL
            FROM disclosure_info
            WHERE CAST("時刻" AS DATE) BETWEEN '{START_DATE}' AND '{END_DATE}'
            GROUP BY 1
        """).df()
        con.close()
    except Exception as e:
        print(f"   [DBエラー]: {e}")
        df_db = pd.DataFrame(columns=['date', 'pdf_db', 'db_XBRL'])
    print(f"   -> DuckDB取得完了 ({time.time() - step_start:.2f}秒)")

    # 3. TDnet Webサイトからの取得
    print("3/5: TDnetWebサイトから件数を取得中 (直近31日分)...")
    step_start = time.time()
    all_dates = pd.date_range(start=START_DATE, end=END_DATE).strftime('%Y-%m-%d').tolist()
    results_map = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(fetch_tdnet_count_single, d, valid_start_date): d for d in all_dates}
        for future in futures:
            d_str, count = future.result()
            results_map[d_str] = count
    print(f"   -> TDnet取得完了 ({time.time() - step_start:.2f}秒)")

    # 4. データのマージ
    print("4/5: データをマージ中...")
    step_start = time.time()
    df_final = pd.DataFrame({"date": all_dates})
    df_final = pd.merge(df_final, df_folder_agg, on="date", how="left")
    df_final = pd.merge(df_final, df_db, on="date", how="left")
    df_final["tdnet_count"] = df_final["date"].map(results_map)
    
    cols_fill = ["pdf_folder", "folder_XBRL", "pdf_db", "db_XBRL", "tdnet_count"]
    for col in cols_fill:
        if col not in df_final.columns: df_final[col] = 0
    df_final[cols_fill] = df_final[cols_fill].fillna(0).astype(int)
    df_final["日付"] = pd.to_datetime(df_final["date"])  # datetimeオブジェクトのまま保持
    print(f"   -> データマージ完了 ({time.time() - step_start:.2f}秒)")

    # 5. ファイル名比較
    print("5/5: ファイル名比較とExcel出力中...")
    step_start = time.time()
    df_filename_comparison = compare_filenames()

    # Excelファイル作成
    wb = Workbook()
    
    # 最初のシートを削除してからシートを作成
    wb.remove(wb.active) # type: ignore
    
    # ファイル数タブの作成
    ws_file_count = wb.create_sheet(TAB_FILE_COUNT)
    ws_file_count.sheet_view.showGridLines = False
    
    headers = [
        "日付", "pdf_folder", "pdf_db", "pdf_folder_DB差異", 
        "folder_XBRL", "db_XBRL", "XBRL数誤差", 
        "TDnet件数", "TDnet_pdf_folderズレ", "TDnet_pdf_dbズレ"
    ]
    ws_file_count.append(headers)

    # スタイル定義
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gray_fill = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
    custom_font = Font(name=FONT_NAME)
    
    side_thin = Side(style='thin', color="000000")
    side_dotted = Side(style='dotted', color="000000")
    
    # 基本の格子（上下）
    border_base = Border(top=side_thin, bottom=side_thin)

    for i, row in df_final.iterrows():
        r_idx = i + 2 # type: ignore
        # 日付をExcelの日付型として設定
        ws_file_count.cell(r_idx, 1, row["日付"])
        ws_file_count.cell(r_idx, 2, int(row["pdf_folder"]))
        ws_file_count.cell(r_idx, 3, int(row["pdf_db"]))
        ws_file_count.cell(r_idx, 4, f"=B{r_idx}-C{r_idx}")
        ws_file_count.cell(r_idx, 5, int(row["folder_XBRL"]))
        ws_file_count.cell(r_idx, 6, int(row["db_XBRL"]))
        ws_file_count.cell(r_idx, 7, f"=E{r_idx}-F{r_idx}")
        ws_file_count.cell(r_idx, 8, int(row["tdnet_count"]))
        ws_file_count.cell(r_idx, 9, f'=IF(H{r_idx}>0, H{r_idx}-B{r_idx}, "")')
        ws_file_count.cell(r_idx, 10, f'=IF(H{r_idx}>0, H{r_idx}-C{r_idx}, "")')

    max_row = ws_file_count.max_row
    max_col = len(headers)
    
    # 罫線と背景色の適用ルール定義
    # A:1, B:2, C:3, D:4, E:5, F:6, G:7, H:8, I:9, J:10
    solid_right_cols = [1, 4, 7, 10]  # 日付, 差異系
    dotted_right_cols = [2, 3, 5, 6, 8, 9] # その他
    gray_cols = [4, 7, 10] # pdf_folder_DB差異, XBRL数誤差, TDnet_pdf_dbズレ

    for r in range(1, 1001):
        for c in range(1, 27):
            cell = ws_file_count.cell(r, c)
            # デフォルト背景
            cell.fill = gray_fill if c in gray_cols else white_fill
            cell.font = custom_font
            
            if r <= max_row and c <= max_col:
                # 罫線の設定
                r_side = side_thin if c in solid_right_cols else side_dotted if c in dotted_right_cols else None
                cell.border = Border(top=side_thin, bottom=side_thin, left=(side_thin if c==1 else None), right=r_side)
                
                if r == 1: # ヘッダー
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name=FONT_NAME, bold=True)
                else:
                    if c == 1: # 日付列
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.number_format = 'yy/mm/dd'  # yy/mm/dd形式に設定
                    else: # 数値列
                        cell.number_format = '#,##0;[Red]-#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in range(1, max_col + 1):
        ws_file_count.column_dimensions[get_column_letter(c)].width = 18

    ws_file_count.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"
    ws_file_count.freeze_panes = "A2"
    
    # ファイル名タブの作成
    ws_filename = wb.create_sheet(TAB_FILENAME_CHECK)
    ws_filename.sheet_view.showGridLines = False
    
    # ヘッダー
    filename_headers = ["種類", "公開日", "連番", "pdfDL", "xbrlDL", "ファイル名", "判定"]
    ws_filename.append(filename_headers)
    
    # データ書き込み
    for i, row in df_filename_comparison.iterrows():
        r_idx = i + 2 # type: ignore
        ws_filename.cell(r_idx, 1, row["種類"])
        ws_filename.cell(r_idx, 2, row["公開日"])
        ws_filename.cell(r_idx, 3, row["連番"])
        ws_filename.cell(r_idx, 4, row["pdfDL"])
        ws_filename.cell(r_idx, 5, row["xbrlDL"])
        ws_filename.cell(r_idx, 6, row["ファイル名"])
        ws_filename.cell(r_idx, 7, row["判定"])
    
    # スタイル適用（ファイル名タブ）
    max_row_filename = ws_filename.max_row
    max_col_filename = len(filename_headers)
    
    for r in range(1, max_row_filename + 1):
        for c in range(1, max_col_filename + 1):
            cell = ws_filename.cell(r, c)
            cell.font = custom_font
            
            if r == 1: # ヘッダー
                cell.fill = gray_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name=FONT_NAME, bold=True)
                cell.border = Border(top=side_thin, bottom=side_thin, 
                                   left=(side_thin if c==1 else None), 
                                   right=side_thin)
            else:
                cell.border = Border(top=side_thin, bottom=side_thin, 
                                   left=(side_thin if c==1 else None), 
                                   right=side_thin)
                if c == 1: # 種類列
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c in [2, 3]: # 公開日、連番列
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else: # その他の列
                    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 列幅設定
    ws_filename.column_dimensions["A"].width = 10  # 種類
    ws_filename.column_dimensions["B"].width = 15  # 公開日
    ws_filename.column_dimensions["C"].width = 12  # 連番
    ws_filename.column_dimensions["D"].width = 15  # pdfDL
    ws_filename.column_dimensions["E"].width = 15  # xbrlDL
    ws_filename.column_dimensions["F"].width = 80  # ファイル名
    ws_filename.column_dimensions["G"].width = 15  # 判定
    
    ws_filename.auto_filter.ref = f"A1:G{max_row_filename}"
    ws_filename.freeze_panes = "A2"

    try:
        wb.save(OUTPUT_EXCEL)
        print(f"   -> Excel出力完了 ({time.time() - step_start:.2f}秒)")
    except PermissionError:
        print(f"\n[エラー] {OUTPUT_EXCEL} を閉じてから再実行してください。")

    end_time = datetime.now()
    print(f"\n" + "="*35)
    print(f"全処理が完了しました。")
    print(f"総実行時間: {time.time() - start_total_time:.2f}秒")
    print(f"終了時刻: {end_time.strftime('%H:%M:%S')}")
    print("="*35)

if __name__ == "__main__":
    process_all_data()
