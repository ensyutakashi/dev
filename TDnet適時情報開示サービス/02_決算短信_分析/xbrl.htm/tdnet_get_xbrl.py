# --- obsidian_property ---
# scr名: 【自動】
# 概要: TDnetの-ixbrl.htmから決算短信データを抽出してDBに保存
# 処理grp: XBRL決算短信取得
# 処理順番: -
# mermaid: "[[mermaid_TDnet_XBRL]]"
# tags: ["tdnet", "決算短信"]
# aliases: ["tdnet_get_xbrl.py"]
# created: 2026-03-26
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：TDnetの-ixbrl.htmから決算短信データを抽出してDBに保存
#　出力ファイルはrawdataとrawdataを加工したCSVファイル
#　加工ファイルのlabel_jpでの日本語名はtse-ed-t-2014-01-12-lab.xmlからデータを抽出したエクセルファイル
#　MAX_FILESで処理ファイル数上限設定
#　■■■変数の設定箇所を確認後実行!!■■■
#
# --- 概要 ---

from __future__ import annotations
import argparse
import csv
import json
import re
import time
from datetime import datetime
import duckdb
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from lxml import etree
from openpyxl import load_workbook
import pandas as pd

# XBRL / iXBRL を読むときに使う代表的な名前空間。
NS = {
    "xhtml": "http://www.w3.org/1999/xhtml",
    "xbrli": "http://www.xbrl.org/2003/instance",
    "xbrldi": "http://xbrl.org/2006/xbrldi",
}

# ==============================
# 🔧 設定（ここだけ変更）
# ==============================
# 入力フォルダ設定
# DEFAULT_INPUT_FILE = Path(r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス\00_TDnet(決算短信)XBRL\test")  # testフォルダ
DEFAULT_INPUT_FILE = Path(r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス\00_TDnet(決算短信)XBRL\2026年")  # 2026年

# フォルダ条件
FOLDER_NAME_FILTER = "_決算短信_"  # フォルダ名に含める文字列（空なら無効）
SUBFOLDER_FILTER = "Summary"  # サブフォルダ名に含める文字列（空なら無効）
# ファイル条件
FILE_NAME_SUFFIX = "-ixbrl.htm"  # 対象ファイル名（この文字列で終わるファイルのみ）
MAX_FILES = 1000  # 処理ファイル数の上限
# 出力先
DEFAULT_OUTPUT_DIR = Path(__file__).parent / "out_db"

# 出力ファイル名設定
OUTPUT_FILES = {
    "facts_csv": "facts_raw.csv",
    "contexts_csv": "contexts_raw.csv", 
    "units_csv": "units_raw.csv",
    "normalized_csv": "facts_normalized.csv",
    "duckdb": "tdnet_financials.duckdb",
    "summary": "summary.json"
}

# label_map.xlsxのパス設定
LABEL_MAP_FILE = Path(r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス\01_tag・value分析\03_lab.xml\label_map(tse-ed-t-2014-01-12-lab.xml).xlsx")
# xbrli_contextRef.xlsxのパス設定（DuckDBに追加する場合）
CONTEXT_REF_FILE = Path(r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス\01_tag・value分析\01_xbrl\context\xbrli_contextRef.xlsx")
# 設定値（引数なしで実行する場合にここを変更）
CONFIG = {
    "input_path": DEFAULT_INPUT_FILE,
    "outdir": DEFAULT_OUTPUT_DIR,
    "metric_map": LABEL_MAP_FILE,
    "context_ref_file": CONTEXT_REF_FILE,
}
#設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定設定

def clean_text(value: Optional[str]) -> str:
    # XML 由来の改行や連続空白を、比較・保存しやすい形に整える。
    if value is None:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def parse_xml(path: Path):
    # 実務データは多少壊れていることがあるので recover=True で読みにいく。
    parser = etree.XMLParser(recover=True, huge_tree=True, remove_blank_text=False)
    return etree.parse(str(path), parser)


def local_name(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def concept_local_name(name_attr: str) -> str:
    return name_attr.split(":", 1)[1] if ":" in name_attr else name_attr


def normalize_numeric(display_value: str, scale: str, sign: str) -> str:
    # 画面表示用の値から、符号と scale を反映した保存向け数値文字列を作る。
    text = clean_text(display_value).replace(",", "")
    if not text:
        return ""
    if sign == "-" and not text.startswith("-"):
        text = "-" + text
    if not scale or scale == "0":
        return text
    try:
        return str(Decimal(text) * (Decimal(10) ** int(scale)))
    except Exception:
        return text


def load_label_map_from_excel(excel_file: Path) -> Dict[str, str]:
    """label_map.xlsxからconcept->日本語ラベルの辞書を作成する"""
    if not excel_file.exists():
        print(f"[警告] label_map.xlsxが見つかりません: {excel_file}")
        return {}
    
    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        label_map = {}
        
        for row in ws.iter_rows(min_row=2, values_only=True):  # 1行目はヘッダー
            if len(row) >= 2 and row[0] and row[1]:
                concept = str(row[0]).strip()
                label_jp = str(row[1]).strip()
                label_map[concept] = label_jp
        
        wb.close()
        return label_map
    except Exception as e:
        print(f"[警告] label_map.xlsxの読み込みに失敗: {e}")
        return {}


def create_metric_map_from_label(label_map: Dict[str, str]) -> Dict[str, str]:
    """日本語ラベルからconcept→metricマッピングを作成する"""
    metric_map = {}
    
    for concept, jp_label in label_map.items():
        # 日本語ラベルからmetric名を推定
        if "売上" in jp_label and "純" in jp_label:
            metric_map[concept] = "net_sales"
        elif "売上" in jp_label and "営業" in jp_label:
            metric_map[concept] = "operating_revenue"
        elif "売上" in jp_label or "収益" in jp_label:
            metric_map[concept] = "sales"
        elif "営業利益" in jp_label:
            metric_map[concept] = "operating_income"
        elif "経常利益" in jp_label:
            metric_map[concept] = "ordinary_income"
        elif "純利益" in jp_label and "親会社" in jp_label:
            metric_map[concept] = "profit_attributable_to_owners"
        elif "純利益" in jp_label:
            metric_map[concept] = "profit"
        elif "1株当たり" in jp_label and "基本" in jp_label:
            metric_map[concept] = "eps_basic"
        elif "1株当たり" in jp_label and "希薄化" in jp_label:
            metric_map[concept] = "eps_diluted"
        elif "資産" in jp_label and "総" in jp_label:
            metric_map[concept] = "total_assets"
        elif "資産" in jp_label and "純" in jp_label:
            metric_map[concept] = "net_assets"
        elif "資本" in jp_label:
            metric_map[concept] = "equity"
        elif "現金" in jp_label and "預金" in jp_label:
            metric_map[concept] = "cash_and_deposits"
        else:
            # conceptを小文字に変換してデフォルト値とする
            metric_map[concept] = concept.lower()
    
    return metric_map


def load_metric_map(path: Optional[Path]) -> Dict[str, str]:
    # label_map.xlsxから動的にマッピングを生成
    if path and path.exists() and path.suffix.lower() == ".xlsx":
        label_map = load_label_map_from_excel(path)
        if label_map:
            metric_map = create_metric_map_from_label(label_map)
            print(f"[INFO] label_map.xlsxから{len(metric_map)}件のマッピングを生成しました")
            return metric_map
    
    # 従来の外部ファイル対応
    if path and path.exists():
        if path.suffix.lower() == ".json":
            data = json.loads(path.read_text(encoding="utf-8"))
        else:
            data = {}
            for line in path.read_text(encoding="utf-8").splitlines():
                s = line.strip()
                if not s or s.startswith("#") or ":" not in s:
                    continue
                k, v = s.split(":", 1)
                data[k.strip()] = v.strip().strip('"').strip("'")
        return data  # 外部ファイルのマッピングのみを返す
    
    # デフォルトは空の辞書を返す
    print(f"[INFO] 外部マッピングファイルがないため、空のマッピングを使用します")
    return {}


def collect_contexts(doc) -> Dict[str, dict]:
    # context は「どの期間・どの連結区分・どの軸の値か」を表す重要情報。
    contexts = {}
    for ctx in doc.xpath("//xbrli:context", namespaces=NS):
        explicit_members = []
        typed_members = []
        for mem in ctx.xpath(".//xbrldi:explicitMember", namespaces=NS):
            explicit_members.append({
                "dimension": mem.get("dimension", ""),
                "value": clean_text("".join(mem.itertext())),
            })
        for mem in ctx.xpath(".//xbrldi:typedMember", namespaces=NS):
            typed_members.append({
                "dimension": mem.get("dimension", ""),
                "value": clean_text("".join(mem.itertext())),
            })

        start_date = clean_text(ctx.findtext("./xbrli:period/xbrli:startDate", namespaces=NS))
        end_date = clean_text(ctx.findtext("./xbrli:period/xbrli:endDate", namespaces=NS))
        instant = clean_text(ctx.findtext("./xbrli:period/xbrli:instant", namespaces=NS))
        # 期間系 context か、一時点系 context かをここで判定しておく。
        if instant:
            period_type = "instant"
        elif start_date or end_date:
            period_type = "duration"
        else:
            period_type = ""

        contexts[ctx.get("id", "")] = {
            "context_id": ctx.get("id", ""),
            "entity_identifier": clean_text(ctx.findtext("./xbrli:entity/xbrli:identifier", namespaces=NS)),
            "period_type": period_type,
            "start_date": start_date,
            "end_date": end_date,
            "instant": instant,
            "explicit_members": explicit_members,
            "typed_members": typed_members,
        }
    return contexts


def collect_units(doc) -> Dict[str, dict]:
    # 単位情報は unit id から fact に結びつけるため、先に辞書化しておく。
    units = {}
    for unit in doc.xpath("//xbrli:unit", namespaces=NS):
        measures = [clean_text(x.text) for x in unit.xpath("./xbrli:measure", namespaces=NS)]
        numerators = [clean_text(x.text) for x in unit.xpath(".//xbrli:unitNumerator/xbrli:measure", namespaces=NS)]
        denominators = [clean_text(x.text) for x in unit.xpath(".//xbrli:unitDenominator/xbrli:measure", namespaces=NS)]
        units[unit.get("id", "")] = {
            "unit_id": unit.get("id", ""),
            "measures": " | ".join(measures),
            "numerators": " | ".join(numerators),
            "denominators": " | ".join(denominators),
        }
    return units


def collect_facts(doc) -> List[dict]:
    # iXBRL 本体から数値/文字列 fact を抜き出し、後段で使いやすい行形式に寄せる。
    rows = []
    ix_uris = [
        "http://www.xbrl.org/2013/inlineXBRL",
        "http://www.xbrl.org/2008/inlineXBRL",
    ]
    idx = 0
    for uri in ix_uris:
        # 2013/2008 の inline XBRL 名前空間の両方に対応する。
        facts = doc.xpath(
            f"//*[namespace-uri()='{uri}' and (local-name()='nonFraction' or local-name()='nonNumeric')]"
        )
        for elem in facts:
            idx += 1
            fact_type = local_name(elem.tag)
            raw_text = clean_text("".join(elem.itertext()))
            is_nil = elem.get("{http://www.w3.org/2001/XMLSchema-instance}nil") == "true"
            rows.append({
                "fact_type": fact_type,
                "name": elem.get("name", ""),
                "contextRef": elem.get("contextRef", ""),
                "unitRef": elem.get("unitRef", ""),
                "decimals": elem.get("decimals", ""),
                "scale": elem.get("scale", ""),
                "sign": elem.get("sign", ""),
                "format": elem.get("format", ""),
                "escape": elem.get("escape", ""),
                "value": "" if is_nil else raw_text,
                "xsi_nil": "true" if is_nil else "",
            })
    return rows


def context_to_dimensions(ctx: dict) -> dict:
    # context から分析しやすい列を派生させる。
    dims = {
        "period_kind": ctx.get("period_type", ""),
        "start_date": ctx.get("start_date", ""),
        "end_date": ctx.get("end_date", ""),
        "instant": ctx.get("instant", ""),
        "consolidation": "",
        "result_type": "",
        "period_scope": "",
        "forecast_kind": "",
        "members_json": json.dumps(ctx.get("explicit_members", []), ensure_ascii=False),
    }

    members = ctx.get("explicit_members", [])
    # 軸メンバーの文字列から、連結区分や実績/予想をざっくり判定する。
    text_blob = " | ".join([m.get("dimension", "") + "=" + m.get("value", "") for m in members]).lower()

    if "consolidatedmember" in text_blob:
        dims["consolidation"] = "consolidated"
    elif "nonconsolidatedmember" in text_blob:
        dims["consolidation"] = "non_consolidated"

    if "revisionforecastmember" in text_blob:
        dims["result_type"] = "forecast_revision"
        dims["forecast_kind"] = "forecast_revision"
    elif "forecastmember" in text_blob:
        dims["result_type"] = "forecast"
        dims["forecast_kind"] = "forecast"
    elif "resultmember" in text_blob:
        dims["result_type"] = "result"

    ctxid = (ctx.get("context_id") or "").lower()
    # context id に含まれる慣例的な命名から、四半期/通期/時点を推定する。
    if "currentyearduration" in ctxid or "currentyear" in ctxid:
        dims["period_scope"] = "full_year"
    elif "currentq1duration" in ctxid or "q1" in ctxid:
        dims["period_scope"] = "q1"
    elif "currentq2duration" in ctxid or "q2" in ctxid or "secondquarter" in ctxid:
        dims["period_scope"] = "q2"
    elif "currentq3duration" in ctxid or "q3" in ctxid or "thirdquarter" in ctxid:
        dims["period_scope"] = "q3"
    elif "instant" in ctxid:
        dims["period_scope"] = "instant"

    return dims


def derive_fiscal_year(ctx: dict) -> str:
    # 終了日または時点日の先頭4桁を会計年度として使う。
    end_date = ctx.get("end_date") or ctx.get("instant") or ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", end_date):
        return end_date[:4]
    return ""


def derive_company_code_from_path(path: Path) -> str:
    # ファイル名中の証券コードらしき連番を拾う。
    m = re.search(r"-(\d{4,6})-", path.name)
    return m.group(1) if m else ""


def derive_doc_type_from_path(path: Path) -> str:
    # ファイル名中のdocファミリー識別子（acedjpsm等）を抽出する。
    name = path.name.lower()
    m = re.search(r"tse-([a-z0-9]+)-", name)
    return m.group(1) if m else ""


def normalize_one(ixbrl_path: Path, metric_map: Dict[str, str]) -> Tuple[List[dict], List[dict], List[dict], List[dict]]:
    # 1ファイル分を、生fact/contexts/unitsテーブル用と正規化テーブル用の4系統で組み立てる。
    doc = parse_xml(ixbrl_path)
    contexts = collect_contexts(doc)
    units = collect_units(doc)
    facts = collect_facts(doc)

    # 日本語ラベルも読み込む
    label_map = load_label_map_from_excel(LABEL_MAP_FILE)

    normalized_rows: List[dict] = []
    facts_rows: List[dict] = []
    contexts_rows: List[dict] = []
    units_rows: List[dict] = []

    company_code = derive_company_code_from_path(ixbrl_path)
    doc_type = derive_doc_type_from_path(ixbrl_path)

    # contextsを行形式に変換
    for ctx_id, ctx in contexts.items():
        # 軸メンバーを個別の列に展開
        annual_dividend = ""
        consolidation = ""
        result_forecast = ""
        
        for member in ctx["explicit_members"]:
            dimension = member.get("dimension", "")
            value = member.get("value", "")
            
            if "AnnualDividendPaymentScheduleAxis" in dimension:
                annual_dividend = value
            elif "ConsolidatedNonconsolidatedAxis" in dimension:
                consolidation = value
            elif "ResultForecastAxis" in dimension:
                result_forecast = value
        
        contexts_rows.append({
            "source_file": ixbrl_path.name,
            "context_id": ctx["context_id"],
            "entity_identifier": ctx["entity_identifier"],
            "period_type": ctx["period_type"],
            "start_date": ctx["start_date"],
            "end_date": ctx["end_date"],
            "instant": ctx["instant"],
            "annual_dividend_payment_schedule": annual_dividend,
            "consolidation_nonconsolidation": consolidation,
            "result_forecast": result_forecast,
            "typed_members_json": json.dumps(ctx["typed_members"], ensure_ascii=False),
        })

    # unitsを行形式に変換
    for unit_id, unit in units.items():
        units_rows.append({
            "source_file": ixbrl_path.name,
            "unit_id": unit["unit_id"],
            "measures": unit["measures"],
            "numerators": unit["numerators"],
            "denominators": unit["denominators"],
        })

    for fact in facts:
        # fact 本体に context / unit の参照情報を結合
        ctx = contexts.get(fact["contextRef"], {})
        unit = units.get(fact["unitRef"], {})
        dims = context_to_dimensions(ctx)
        concept = concept_local_name(fact["name"]) if fact["name"] else ""
        metric = metric_map.get(concept, concept)
        label_jp = label_map.get(concept, "")  # 日本語ラベルを取得

        facts_rows.append({
            "source_file": ixbrl_path.name,
            "doc_type": doc_type,
            "code": company_code,
            **fact,
        })

        normalized_rows.append({
            "source_file": ixbrl_path.name,
            "doc_type": doc_type,
            "code": company_code,
            "fiscal_year": derive_fiscal_year(ctx),
            "metric": metric,
            "concept": concept,
            "label_jp": label_jp,  # 日本語ラベルを追加
            "fact_type": fact["fact_type"],
            "value": fact["value"],
            "unit": unit.get("measures", ""),
            "unit_numerators": unit.get("numerators", ""),
            "unit_denominators": unit.get("denominators", ""),
            "period_kind": dims["period_kind"],
            "period_scope": dims["period_scope"],
            "start_date": dims["start_date"],
            "end_date": dims["end_date"],
            "instant": dims["instant"],
            "consolidation": dims["consolidation"],
            "result_type": dims["result_type"],
            "forecast_kind": dims["forecast_kind"],
            "members_json": dims["members_json"],
            "context_id": ctx.get("context_id", ""),
            "context_ref": fact["contextRef"],
            "unit_ref": fact["unitRef"],
        })

    return facts_rows, contexts_rows, units_rows, normalized_rows


def write_csv(path: Path, rows: List[dict]) -> None:
    # 空でもヘッダなしの空ファイルとして出力し、後続処理で扱いやすくする。
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    
    # contexts_csvの場合は列順を指定
    if path.name == "contexts_raw.csv":
        fieldnames = [
            "source_file", "context_id", "entity_identifier", "period_type", 
            "start_date", "end_date", "instant",
            "annual_dividend_payment_schedule", "consolidation_nonconsolidation", "result_forecast",
            "typed_members_json"
        ]
    else:
        fieldnames = list(rows[0].keys())
    
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_context_ref_from_excel(excel_file: Path) -> List[dict]:
    """xbrli_contextRef.xlsxからcontext情報を読み込む"""
    if not excel_file.exists():
        print(f"[警告] {excel_file.name}が見つかりません: {excel_file}")
        return []
    
    try:
        df = pd.read_excel(excel_file)
        # カラム名を小文字に変換して統一
        df.columns = [col.lower().replace(' ', '_') for col in df.columns]
        return df.to_dict('records')
    except Exception as e:
        print(f"[警告] {excel_file.name}の読み込みに失敗: {e}")
        return []


def write_duckdb(path: Path, facts_rows: List[dict], contexts_rows: List[dict], units_rows: List[dict], normalized_rows: List[dict], context_ref_rows: List[dict] = None) -> None:
    # 毎回作り直す前提なので、既存DBがあれば削除して再生成する。
    if path.exists():
        try:
            path.unlink()
        except PermissionError:
            print(f"[警告] DBファイルが使用中のため削除できません。上書きで続行します: {path}")
    conn = duckdb.connect(str(path))
    cur = conn.cursor()

    # 既存テーブルを削除
    cur.execute("DROP TABLE IF EXISTS facts_raw")
    cur.execute("DROP TABLE IF EXISTS contexts_raw")
    cur.execute("DROP TABLE IF EXISTS units_raw")
    cur.execute("DROP TABLE IF EXISTS facts_normalized")
    if context_ref_rows:
        cur.execute("DROP TABLE IF EXISTS context_ref")

    # facts_rawテーブル
    cur.execute("""
        CREATE TABLE facts_raw (
            source_file VARCHAR,
            doc_type VARCHAR,
            code VARCHAR,
            fact_type VARCHAR,
            name VARCHAR,
            contextRef VARCHAR,
            unitRef VARCHAR,
            decimals VARCHAR,
            scale VARCHAR,
            sign VARCHAR,
            format VARCHAR,
            escape VARCHAR,
            value VARCHAR,
            xsi_nil VARCHAR
        )
    """)

    # contexts_rawテーブル
    cur.execute("""
        CREATE TABLE contexts_raw (
            source_file VARCHAR,
            context_id VARCHAR,
            entity_identifier VARCHAR,
            period_type VARCHAR,
            start_date VARCHAR,
            end_date VARCHAR,
            instant VARCHAR,
            annual_dividend_payment_schedule VARCHAR,
            consolidation_nonconsolidation VARCHAR,
            result_forecast VARCHAR,
            typed_members_json VARCHAR,
            PRIMARY KEY (source_file, context_id)
        )
    """)

    # units_rawテーブル
    cur.execute("""
        CREATE TABLE units_raw (
            source_file VARCHAR,
            unit_id VARCHAR,
            measures VARCHAR,
            numerators VARCHAR,
            denominators VARCHAR,
            PRIMARY KEY (source_file, unit_id)
        )
    """)

    # facts_normalizedテーブル
    cur.execute("""
        CREATE TABLE facts_normalized (
            source_file VARCHAR,
            doc_type VARCHAR,
            code VARCHAR,
            fiscal_year VARCHAR,
            metric VARCHAR,
            concept VARCHAR,
            label_jp VARCHAR,
            fact_type VARCHAR,
            value VARCHAR,
            unit VARCHAR,
            unit_numerators VARCHAR,
            unit_denominators VARCHAR,
            period_kind VARCHAR,
            period_scope VARCHAR,
            start_date VARCHAR,
            end_date VARCHAR,
            instant VARCHAR,
            consolidation VARCHAR,
            result_type VARCHAR,
            forecast_kind VARCHAR,
            members_json VARCHAR,
            context_id VARCHAR,
            context_ref VARCHAR,
            unit_ref VARCHAR
        )
    """)

    # データ挿入
    if facts_rows:
        cols = list(facts_rows[0].keys())
        placeholders = ','.join(['?'] * len(cols))
        cur.executemany(
            f"INSERT INTO facts_raw ({','.join(cols)}) VALUES ({placeholders})",
            [tuple(r[col] for col in cols) for r in facts_rows]
        )

    if contexts_rows:
        cols = list(contexts_rows[0].keys())
        placeholders = ','.join(['?'] * len(cols))
        cur.executemany(
            f"INSERT INTO contexts_raw ({','.join(cols)}) VALUES ({placeholders})",
            [tuple(r[col] for col in cols) for r in contexts_rows]
        )

    if units_rows:
        cols = list(units_rows[0].keys())
        placeholders = ','.join(['?'] * len(cols))
        cur.executemany(
            f"INSERT INTO units_raw ({','.join(cols)}) VALUES ({placeholders})",
            [tuple(r[col] for col in cols) for r in units_rows]
        )

    if normalized_rows:
        cols = list(normalized_rows[0].keys())
        placeholders = ','.join(['?'] * len(cols))
        cur.executemany(
            f"INSERT INTO facts_normalized ({','.join(cols)}) VALUES ({placeholders})",
            [tuple(r[col] for col in cols) for r in normalized_rows]
        )

    # context_refテーブルを作成してデータを挿入
    if context_ref_rows:
        # カラム名を取得してテーブル作成
        cols = list(context_ref_rows[0].keys())
        col_definitions = ', '.join([f"{col} VARCHAR" for col in cols])
        cur.execute(f"CREATE TABLE context_ref ({col_definitions})")
        
        # データ挿入
        placeholders = ','.join(['?'] * len(cols))
        cur.executemany(
            f"INSERT INTO context_ref ({','.join(cols)}) VALUES ({placeholders})",
            [tuple(r[col] for col in cols) for r in context_ref_rows]
        )
        print(f"[INFO] context_refテーブルに{len(context_ref_rows)}件のデータを挿入しました")

    # インデックス作成
    cur.execute("CREATE INDEX idx_norm_company_metric ON facts_normalized(code, metric)")
    cur.execute("CREATE INDEX idx_norm_period ON facts_normalized(fiscal_year, period_scope, result_type)")
    cur.execute("CREATE INDEX idx_facts_context_ref ON facts_raw(contextRef)")
    cur.execute("CREATE INDEX idx_facts_unit_ref ON facts_raw(unitRef)")
    if context_ref_rows:
        # 実際のカラム名を確認してインデックスを作成
        cols = list(context_ref_rows[0].keys())
        if 'context_id' in cols:
            cur.execute("CREATE INDEX idx_context_ref_id ON context_ref(context_id)")
        elif 'contextref' in cols:
            cur.execute("CREATE INDEX idx_context_ref_id ON context_ref(contextref)")
        else:
            print(f"[警告] context_refテーブルにインデックス用のカラムが見つかりません。利用可能なカラム: {cols}")
    conn.commit()
    conn.close()


def find_ixbrl_files(input_path: Path) -> List[Path]:
    # 単一ファイル指定にも、ディレクトリ一括処理にも対応する。
    if input_path.is_file():
        return [input_path]
    
    candidates = []
    
    for p in input_path.rglob("*"):
        low = p.name.lower()
        
        # 条件1: ixbrlを含むhtmファイル
        if not (p.is_file() and low.endswith((".htm", ".html", ".xhtml")) and "ixbrl" in low):
            continue
        
        # 条件2: 設定されたフォルダ名フィルタ
        if FOLDER_NAME_FILTER and FOLDER_NAME_FILTER not in str(p.parent):
            continue
        
        # 条件3: 設定されたサブフォルダ名フィルタ
        if SUBFOLDER_FILTER and SUBFOLDER_FILTER not in str(p.parent):
            continue
        
        # 条件4: 設定されたファイル名サフィックス
        if FILE_NAME_SUFFIX and not p.name.endswith(FILE_NAME_SUFFIX):
            continue
        
        candidates.append(p)
        
        # 上限チェック
        if MAX_FILES and len(candidates) >= MAX_FILES:
            print(f"[INFO] 処理ファイル数が上限({MAX_FILES})に達したため、残りはスキップします")
            break
    
    print(f"[INFO] {len(candidates)}件のiXBRLファイルを検出しました")
    return sorted(candidates)


def main():
    # CLI から対象ファイル群を読み、CSV と DuckDB をまとめて出力する。
    start_time = datetime.now()
    start_perf = time.perf_counter()
    print(f"[INFO] 開始時間: {start_time:%Y-%m-%d %H:%M:%S}")

    result_code = 0
    try:
        ap = argparse.ArgumentParser(description="TDnet iXBRL -> 正規化DB/CSV")
        ap.add_argument("input", type=Path, nargs="?", help="ixbrl file or folder (省略時は設定値を使用)")
        ap.add_argument("-o", "--outdir", type=Path, default=None, help="output directory")
        ap.add_argument("--metric-map", type=Path, default=None, help="json or yaml-like key:value file (省略時はスクリプト冒頭のLABEL_MAP_FILEを使用)")
        args = ap.parse_args()

        # 引数がなければ設定値を使用
        if args.input is None:
            input_path = CONFIG["input_path"]
            if not input_path.exists():
                print(f"エラー: デフォルト入力ファイルが見つかりません: {input_path}")
                return 1
        else:
            input_path = args.input.resolve()

        outdir = args.outdir.resolve() if args.outdir else CONFIG["outdir"]

        ixbrl_files = find_ixbrl_files(input_path)
        if not ixbrl_files:
            raise SystemExit("ixbrl file not found")

        # metric_mapを読み込む
        if args.metric_map:
            metric_map_path = args.metric_map
        else:
            metric_map_path = CONFIG.get("metric_map")

        if metric_map_path and metric_map_path.exists():
            print(f"[INFO] label_map.xlsxを使用: {metric_map_path}")
            metric_map = load_metric_map(metric_map_path)
        else:
            print("[INFO] label_map.xlsxが見つからないため、デフォルトマッピングを使用します")
            metric_map = {}

        outdir.mkdir(parents=True, exist_ok=True)

        print(f"処理ファイル:")
        print(f"  入力:   {input_path}")
        print(f"  Label:  {metric_map_path if metric_map_path and metric_map_path.exists() else 'なし'}")
        print(f"  出力:   {outdir}")

        all_facts: List[dict] = []
        all_contexts: List[dict] = []
        all_units: List[dict] = []
        all_norm: List[dict] = []
        
        # context_ref.xlsxからデータを読み込む
        context_ref_path = CONFIG.get("context_ref_file")
        context_ref_data = []
        if context_ref_path and context_ref_path.exists():
            print(f"[INFO] context_ref.xlsxを読み込みます: {context_ref_path}")
            context_ref_data = load_context_ref_from_excel(context_ref_path)
        else:
            print("[INFO] context_ref.xlsxが見つからないか、指定されていません")

        for fp in ixbrl_files:
            # ファイル単位で正規化し、最後に全件をまとめて書き出す。
            facts_rows, contexts_rows, units_rows, norm_rows = normalize_one(fp, metric_map)
            all_facts.extend(facts_rows)
            all_contexts.extend(contexts_rows)
            all_units.extend(units_rows)
            all_norm.extend(norm_rows)

        write_csv(outdir / OUTPUT_FILES["facts_csv"], all_facts)
        write_csv(outdir / OUTPUT_FILES["contexts_csv"], all_contexts)
        write_csv(outdir / OUTPUT_FILES["units_csv"], all_units)
        write_csv(outdir / OUTPUT_FILES["normalized_csv"], all_norm)
        write_duckdb(outdir / OUTPUT_FILES["duckdb"], all_facts, all_contexts, all_units, all_norm, context_ref_data)

        summary = {
            "input_count": len(ixbrl_files),
            "facts_count": len(all_facts),
            "contexts_count": len(all_contexts),
            "units_count": len(all_units),
            "normalized_fact_count": len(all_norm),
            "context_ref_count": len(context_ref_data),
            "outdir": str(outdir.resolve()),
            "files": [p.name for p in ixbrl_files],
        }
        (outdir / OUTPUT_FILES["summary"]).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0
    finally:
        end_time = datetime.now()
        elapsed_sec = time.perf_counter() - start_perf
        print(f"[INFO] 終了時間: {end_time:%Y-%m-%d %H:%M:%S}")
        print(f"[INFO] 経過時間: {elapsed_sec:.2f}秒")


if __name__ == "__main__":
    main()
