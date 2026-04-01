# --- obsidian_property ---
# scr名: 【自動】
# 概要: TDnetの決算短信iXBRLファイルから完全な財務データを抽出し、Excelにまとめる
# 処理grp: XBRL
# 処理順番: -
# input: XBRL summary
# output: OUTPUT_FILE
# mermaid: 
# tags: ["XBRL", "python"]
# aliases: 
# created: 2026-03-23
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：TDnetの決算短信iXBRLファイルから完全な財務データを抽出し、Excelにまとめる
#
# 特徴
# 完全なデータ抽出: 値、単位、期間、型、階層など全情報を取得
# 複数シート出力: データ種類ごとに整理されたExcelファイルを生成
# 日本語対応: label_map.xlsx連携で日本語ラベルを表示
#
# 入力ファイル
# iXBRL (.htm) - 実際の財務データ
# XSD (.xsd) - 要素定義
# def.xml (.xml) - 階層関係
# label_map.xlsx - 日本語ラベル辞書
#
# シート一覧:
# ┌─────────────────────┬─────────────────────┬─────────────────┬─────────┐
# │ シート名            │ 意味                 │ 主な用途         │ 重要度  │
# ├─────────────────────┼─────────────────────┼─────────────────┼─────────┤
# │ facts_enriched      │ 日本語ラベル付き主要データ │ 通常の財務確認  │ ★★★    │
# │ hierarchy           │ 階層構造表示          │ 階層の視覚化     │ ★★★    │
# │ summary             │ 全シート概要          │ ファイル全体把握 │ ★★★    │
# │ contexts            │ 期間情報              │ いつのデータか  │ ★★☆    │
# │ units               │ 単位情報              │ 何の単位か      │ ★★☆    │
# │ def_arcs            │ 階層関係定義          │ 階層の技術確認  │ ★★☆    │
# │ facts_all           │ 生データ全て          │ 技術的詳細確認  │ ★☆☆    │
# │ xsd_elements        │ 要素型定義            │ 要素の技術仕様  │ ★☆☆    │
# │ labels              │ 日本語ラベル辞書      │ ラベル参照・修正│ ★☆☆    │
# │ schema_refs         │ スキーマ参照情報      │ 技術的トレーサビリティ│ ★☆☆ │
# │ xsd_linkbase_refs   │ XSDリンクベース参照   │ 技術的参照確認  │ ☆☆☆    │
# │ def_locators        │ 定義位置情報          │ 技術的位置確認  │ ☆☆☆    │
# │ concepts_without_facts│ 値なしの定義済みconcept│ 未使用項目確認 │ ☆☆☆    │
# └─────────────────────┴─────────────────────┴─────────────────┴─────────┘
#
# 使い分け:
# 日常利用: facts_enriched + hierarchy
# 詳細分析: + contexts + units
# 技術確認: + def_arcs + xsd_elements

# --- 概要 ---

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =========================
# 設定：ここでファイル名を指定
# =========================
# スクリプトと同じフォルダに全てのファイルを配置する場合
BASE_DIR = Path(__file__).parent

# 処理対象ファイル（相対パス）
IXBRL_FILE = "tse-acedjpsm-71720-20260209552878-ixbrl.htm"
XSD_FILE = "tse-acedjpsm-71720-20260209552878.xsd"
DEFXML_FILE = "tse-acedjpsm-71720-20260209552878-def.xml"
LABEL_MAP_FILE = "label_map(tse-ed-t-2014-01-12-lab.xml).xlsx"

# 出力ファイル名（スクリプト名_IXBRL_FILEの形式で指定）
OUTPUT_FILE = Path(__file__).stem + "_" + Path(IXBRL_FILE).stem + ".xlsx"

# =========================


IX_URIS = [
    "http://www.xbrl.org/2008/inlineXBRL",
    "http://www.xbrl.org/2013/inlineXBRL",
]

NS = {
    "xhtml": "http://www.w3.org/1999/xhtml",
    "link": "http://www.xbrl.org/2003/linkbase",
    "xlink": "http://www.w3.org/1999/xlink",
    "xbrli": "http://www.xbrl.org/2003/instance",
    "xbrldi": "http://xbrl.org/2006/xbrldi",
    "xs": "http://www.w3.org/2001/XMLSchema",
    "xsd": "http://www.w3.org/2001/XMLSchema",
}


def clean_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def parse_xml(path: Path):
    parser = etree.XMLParser(recover=True, huge_tree=True, remove_blank_text=False)
    return etree.parse(str(path), parser)


def local_name(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def concept_local_name(name_attr: str) -> str:
    return name_attr.split(":", 1)[1] if ":" in name_attr else name_attr


def inner_xml(elem) -> str:
    return etree.tostring(elem, encoding="unicode", with_tail=False)


def element_text(elem) -> str:
    text = "".join(elem.itertext())
    return clean_text(text)


def format_bool_from_ixt(fmt: str, text: str) -> str:
    if text:
        return text
    fmt = (fmt or "").lower()
    if "booleantrue" in fmt:
        return "true"
    if "booleanfalse" in fmt:
        return "false"
    return text


def normalized_numeric(display_value: str, scale: str, sign: str) -> str:
    text = clean_text(display_value).replace(",", "")
    if not text:
        return ""
    if sign == "-" and not text.startswith("-"):
        text = "-" + text
    if not scale or scale == "0":
        return text
    try:
        return str(Decimal(text) * (Decimal(10) ** int(scale)))
    except Exception:
        return text


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths[cell.column], min(len(value), 80))
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = max(10, min(width + 2, 60))


def write_rows(ws, rows: List[Dict[str, object]]):
    if not rows:
        ws.append(["message"])
        ws.append(["no data"])
        style_sheet(ws)
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)
    ws.auto_filter.ref = ws.dimensions


def load_label_map_from_excel(path: Optional[Path]) -> Dict[str, str]:
    if not path or not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    if "label_map" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["label_map"]
    result: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        concept = row[0]
        label = row[1]
        role = row[2] if len(row) > 2 else None
        if not concept or not label:
            continue
        if concept not in result or (role and str(role).endswith("/label")):
            result[str(concept)] = str(label)
    wb.close()
    return result


def collect_ix_facts(doc) -> List[Dict[str, object]]:
    rows = []
    hidden_roots = set()
    for uri in IX_URIS:
        hidden_roots.update(doc.xpath(f"//*[namespace-uri()='{uri}' and local-name()='hidden']"))
    hidden_fact_ids = {id(x) for hidden in hidden_roots for x in hidden.iterdescendants() if local_name(x.tag) in {"nonFraction", "nonNumeric"}}

    fact_index = 0
    for uri in IX_URIS:
        facts = doc.xpath(f"//*[namespace-uri()='{uri}' and (local-name()='nonFraction' or local-name()='nonNumeric')]")
        for elem in facts:
            fact_index += 1
            fact_type = local_name(elem.tag)
            raw_text = element_text(elem)
            raw_text = format_bool_from_ixt(elem.get("format", ""), raw_text)
            is_nil = elem.get("{http://www.w3.org/2001/XMLSchema-instance}nil") == "true"
            row = {
                "fact_index": fact_index,
                "fact_type": fact_type,
                "is_hidden": "Y" if id(elem) in hidden_fact_ids else "",
                "concept_qname": elem.get("name", ""),
                "concept": concept_local_name(elem.get("name", "")) if elem.get("name") else "",
                "contextRef": elem.get("contextRef", ""),
                "unitRef": elem.get("unitRef", ""),
                "decimals": elem.get("decimals", ""),
                "scale": elem.get("scale", ""),
                "sign": elem.get("sign", ""),
                "format": elem.get("format", ""),
                "escape": elem.get("escape", ""),
                "tupleRef": elem.get("tupleRef", ""),
                "order": elem.get("order", ""),
                "continuedAt": elem.get("continuedAt", ""),
                "xml_lang": elem.get("{http://www.w3.org/XML/1998/namespace}lang", ""),
                "xsi_nil": "true" if is_nil else "",
                "display_value": "" if is_nil else raw_text,
                "normalized_value": "" if is_nil else (normalized_numeric(raw_text, elem.get("scale", ""), elem.get("sign", "")) if fact_type == "nonFraction" else raw_text),
                "xpath": elem.getroottree().getpath(elem),
                "raw_xml": inner_xml(elem),
            }
            rows.append(row)
    return rows


def collect_contexts(doc) -> List[Dict[str, object]]:
    rows = []
    for ctx in doc.xpath("//xbrli:context", namespaces=NS):
        explicit_members = []
        typed_members = []
        for mem in ctx.xpath(".//xbrldi:explicitMember", namespaces=NS):
            explicit_members.append(f'{mem.get("dimension","")}={clean_text("".join(mem.itertext()))}')
        for mem in ctx.xpath(".//xbrldi:typedMember", namespaces=NS):
            typed_members.append(f'{mem.get("dimension","")}={clean_text("".join(mem.itertext()))}')

        period_type = ""
        start_date = clean_text(ctx.findtext("./xbrli:period/xbrli:startDate", namespaces=NS))
        end_date = clean_text(ctx.findtext("./xbrli:period/xbrli:endDate", namespaces=NS))
        instant = clean_text(ctx.findtext("./xbrli:period/xbrli:instant", namespaces=NS))
        forever = ctx.find("./xbrli:period/xbrli:forever", namespaces=NS)
        if instant:
            period_type = "instant"
        elif start_date or end_date:
            period_type = "duration"
        elif forever is not None:
            period_type = "forever"

        rows.append({
            "context_id": ctx.get("id", ""),
            "entity_identifier": clean_text(ctx.findtext("./xbrli:entity/xbrli:identifier", namespaces=NS)),
            "entity_scheme": (ctx.find("./xbrli:entity/xbrli:identifier", namespaces=NS).get("scheme", "") if ctx.find("./xbrli:entity/xbrli:identifier", namespaces=NS) is not None else ""),
            "period_type": period_type,
            "start_date": start_date,
            "end_date": end_date,
            "instant": instant,
            "explicit_members": " | ".join(explicit_members),
            "typed_members": " | ".join(typed_members),
            "raw_xml": inner_xml(ctx),
        })
    return rows


def collect_units(doc) -> List[Dict[str, object]]:
    rows = []
    for unit in doc.xpath("//xbrli:unit", namespaces=NS):
        measures = [clean_text(x.text) for x in unit.xpath("./xbrli:measure", namespaces=NS)]
        numerators = [clean_text(x.text) for x in unit.xpath(".//xbrli:unitNumerator/xbrli:measure", namespaces=NS)]
        denominators = [clean_text(x.text) for x in unit.xpath(".//xbrli:unitDenominator/xbrli:measure", namespaces=NS)]
        rows.append({
            "unit_id": unit.get("id", ""),
            "measures": " | ".join(measures),
            "numerators": " | ".join(numerators),
            "denominators": " | ".join(denominators),
            "raw_xml": inner_xml(unit),
        })
    return rows


def collect_schema_refs_from_ix(doc) -> List[Dict[str, object]]:
    rows = []
    for elem in doc.xpath("//link:schemaRef", namespaces=NS):
        rows.append({
            "href": elem.get("{http://www.w3.org/1999/xlink}href", ""),
            "type": elem.get("{http://www.w3.org/1999/xlink}type", ""),
        })
    return rows


def collect_xsd_metadata(xsd_doc) -> Tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    elements = []
    linkbase_refs = []
    for elem in xsd_doc.xpath("//xs:element | //xsd:element", namespaces=NS):
        elements.append({
            "concept": elem.get("name", ""),
            "id": elem.get("id", ""),
            "type": elem.get("type", ""),
            "substitutionGroup": elem.get("substitutionGroup", ""),
            "abstract": elem.get("abstract", ""),
            "nillable": elem.get("nillable", ""),
            "periodType": elem.get("{http://www.xbrl.org/2003/instance}periodType", ""),
            "balance": elem.get("{http://www.xbrl.org/2003/instance}balance", ""),
        })
    for ref in xsd_doc.xpath("//link:linkbaseRef", namespaces=NS):
        linkbase_refs.append({
            "href": ref.get("{http://www.w3.org/1999/xlink}href", ""),
            "arcrole": ref.get("{http://www.w3.org/1999/xlink}arcrole", ""),
            "role": ref.get("{http://www.w3.org/1999/xlink}role", ""),
            "type": ref.get("{http://www.w3.org/1999/xlink}type", ""),
        })
    return elements, linkbase_refs


def concept_from_href(href: str) -> str:
    tail = href.split("#")[-1]
    return tail.split("_", 1)[1] if "_" in tail else tail


def collect_def_data(def_doc) -> Tuple[List[Dict[str, object]], List[Dict[str, object]]]:
    loc_rows = []
    arc_rows = []

    for dlink in def_doc.xpath("//link:definitionLink", namespaces=NS):
        role = dlink.get("{http://www.w3.org/1999/xlink}role", "")
        loc_map: Dict[str, Dict[str, str]] = {}
        for loc in dlink.xpath("./link:loc", namespaces=NS):
            label = loc.get("{http://www.w3.org/1999/xlink}label", "")
            href = loc.get("{http://www.w3.org/1999/xlink}href", "")
            concept = concept_from_href(href)
            loc_map[label] = {"concept": concept, "href": href}
            loc_rows.append({
                "role": role,
                "locator_label": label,
                "concept": concept,
                "href": href,
            })

        for arc in dlink.xpath("./link:definitionArc", namespaces=NS):
            frm = arc.get("{http://www.w3.org/1999/xlink}from", "")
            to = arc.get("{http://www.w3.org/1999/xlink}to", "")
            arc_rows.append({
                "role": role,
                "arcrole": arc.get("{http://www.w3.org/1999/xlink}arcrole", ""),
                "from_locator": frm,
                "from_concept": loc_map.get(frm, {}).get("concept", frm),
                "to_locator": to,
                "to_concept": loc_map.get(to, {}).get("concept", to),
                "order": arc.get("order", ""),
                "priority": arc.get("priority", ""),
                "xbrldt_closed": arc.get("{http://xbrl.org/2005/xbrldt}closed", ""),
                "xbrldt_contextElement": arc.get("{http://xbrl.org/2005/xbrldt}contextElement", ""),
                "raw_xml": inner_xml(arc),
            })
    return loc_rows, arc_rows


def build_fact_enriched(facts, contexts, units, labels) -> List[Dict[str, object]]:
    context_map = {x["context_id"]: x for x in contexts}
    unit_map = {x["unit_id"]: x for x in units}
    rows = []
    for fact in facts:
        ctx = context_map.get(fact["contextRef"], {})
        unit = unit_map.get(fact["unitRef"], {})
        rows.append({
            "fact_index": fact["fact_index"],
            "concept": fact["concept"],
            "label_jp": labels.get(fact["concept"], ""),
            "fact_type": fact["fact_type"],
            "is_hidden": fact["is_hidden"],
            "display_value": fact["display_value"],
            "normalized_value": fact["normalized_value"],
            "contextRef": fact["contextRef"],
            "period_type": ctx.get("period_type", ""),
            "start_date": ctx.get("start_date", ""),
            "end_date": ctx.get("end_date", ""),
            "instant": ctx.get("instant", ""),
            "explicit_members": ctx.get("explicit_members", ""),
            "typed_members": ctx.get("typed_members", ""),
            "unitRef": fact["unitRef"],
            "unit_measures": unit.get("measures", ""),
            "unit_numerators": unit.get("numerators", ""),
            "unit_denominators": unit.get("denominators", ""),
            "decimals": fact["decimals"],
            "scale": fact["scale"],
            "sign": fact["sign"],
            "format": fact["format"],
            "xsi_nil": fact["xsi_nil"],
            "concept_qname": fact["concept_qname"],
        })
    return rows


def build_hierarchy_sheet(facts: List[Dict[str, object]], def_arcs: List[Dict[str, object]], label_map: Dict[str, str]) -> List[Dict[str, object]]:
    """階層表示シートを構築する（summary_to_excel2.pyの機能を統合）"""
    # factsから値の辞書を作成
    values = {}
    for fact in facts:
        concept = fact.get("concept", "")
        value = fact.get("display_value", "")
        if concept and value:
            values[concept] = value
    
    # def_arcsから階層構造を構築
    children_map: Dict[str, List[dict]] = defaultdict(list)
    parent_map: Dict[str, str] = {}
    all_nodes: set = set()
    
    for arc in def_arcs:
        parent = arc.get("from_concept", "")
        child = arc.get("to_concept", "")
        order = arc.get("order", "")
        
        # from_concept/to_conceptが空の場合はfrom_locator/to_locatorを使用
        if not parent:
            parent = arc.get("from_locator", "")
        if not child:
            child = arc.get("to_locator", "")
        
        if parent and child:
            children_map[parent].append({"child": child, "order": order})
            parent_map[child] = parent
            all_nodes.add(parent)
            all_nodes.add(child)
    
    # ルートノードを探す
    roots = [node for node in all_nodes if node not in parent_map]
    roots.sort(key=lambda x: (0 if x == "DocumentEntityInformationHeading" else 1, x))
    
    # 階層展開
    def expand_tree(node: str, level: int, parent_path: List[str], rows: List[dict]):
        current_path = parent_path + [node]
        
        row_data = {
            "Level": level,
            "order": "",
            "値": values.get(node, ""),
        }
        
        # Lv0-Lv5の英語・日本語ラベルを設定
        for i in range(6):
            if i < len(current_path):
                row_data[f"Lv{i}_en"] = current_path[i]
                row_data[f"Lv{i}_jp"] = label_map.get(current_path[i], current_path[i])
            else:
                row_data[f"Lv{i}_en"] = ""
                row_data[f"Lv{i}_jp"] = ""
        
        rows.append(row_data)
        
        # 子ノードを再帰的に展開
        for child_arc in sorted(children_map.get(node, []), key=lambda x: x["order"]):
            expand_tree(child_arc["child"], level + 1, current_path, rows)
    
    # 階層行データを生成
    hierarchy_rows: List[dict] = []
    for root in roots:
        expand_tree(root, 0, [], hierarchy_rows)
    
    return hierarchy_rows


def build_concepts_without_facts(xsd_elements, facts, labels) -> List[Dict[str, object]]:
    fact_concepts = {x["concept"] for x in facts if x.get("concept")}
    rows = []
    for row in xsd_elements:
        concept = row.get("concept", "")
        if concept and concept not in fact_concepts:
            rows.append({
                "concept": concept,
                "label_jp": labels.get(concept, ""),
                "type": row.get("type", ""),
                "periodType": row.get("periodType", ""),
                "balance": row.get("balance", ""),
                "abstract": row.get("abstract", ""),
            })
    return rows


def build_labels_sheet(label_map: Dict[str, str]) -> List[Dict[str, object]]:
    return [{"concept": k, "label_jp": v} for k, v in sorted(label_map.items())]


def build_workbook(output_path: Path, sheets: Dict[str, List[Dict[str, object]]], issues: List[str]):
    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("summary")
    summary_rows = [{"sheet_name": name, "rows": len(rows)} for name, rows in sheets.items()]
    if issues:
        for i, issue in enumerate(issues, 1):
            summary_rows.append({"sheet_name": f"issue_{i}", "rows": issue})
    write_rows(summary_ws, summary_rows)

    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        write_rows(ws, rows)
        
        # 折り返し表示をなくす
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=False)

    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser(description="TDnet iXBRL 全件抽出 -> Excel")
    ap.add_argument("ixbrl", type=Path, nargs="?", help="iXBRL .htm/.html (省略時は設定値を使用)")
    ap.add_argument("--xsd", type=Path, help="instance xsd (省略時は設定値を使用)")
    ap.add_argument("--defxml", type=Path, help="definition linkbase xml (省略時は設定値を使用)")
    ap.add_argument("--label-map-xlsx", type=Path, help="label_map.xlsx (省略時は設定値を使用)")
    ap.add_argument("-o", "--output", type=Path, help="output xlsx (省略時は設定値を使用)")
    args = ap.parse_args()

    # 引数がなければ設定値を使用
    if args.ixbrl is None:
        print("コマンドライン引数がありません。設定値を使用します。")
        ix_path = BASE_DIR / IXBRL_FILE
        xsd_path = BASE_DIR / XSD_FILE
        def_path = BASE_DIR / DEFXML_FILE
        label_map_path = BASE_DIR / LABEL_MAP_FILE if LABEL_MAP_FILE else None
        out_path = BASE_DIR / (OUTPUT_FILE if OUTPUT_FILE else Path(__file__).stem + "_" + Path(IXBRL_FILE).stem + ".xlsx")
    else:
        ix_path = args.ixbrl.resolve()
        xsd_path = args.xsd.resolve() if args.xsd else BASE_DIR / XSD_FILE
        def_path = args.defxml.resolve() if args.defxml else BASE_DIR / DEFXML_FILE
        label_map_path = args.label_map_xlsx.resolve() if args.label_map_xlsx else (BASE_DIR / LABEL_MAP_FILE if LABEL_MAP_FILE else None)
        out_path = args.output.resolve() if args.output else ix_path.with_name(Path(__file__).stem + "_" + ix_path.stem + ".xlsx")

    # ファイル存在チェック
    if not ix_path.exists():
        print(f"エラー: iXBRLファイルが見つかりません: {ix_path}")
        return 1
    if not xsd_path.exists():
        print(f"エラー: XSDファイルが見つかりません: {xsd_path}")
        return 1
    if not def_path.exists():
        print(f"エラー: DefXMLファイルが見つかりません: {def_path}")
        return 1
    if label_map_path and not label_map_path.exists():
        print(f"警告: LabelMapファイルが見つかりません: {label_map_path}")
        label_map_path = None

    print(f"処理ファイル:")
    print(f"  iXBRL: {ix_path}")
    print(f"  XSD:   {xsd_path}")
    print(f"  DefXML:{def_path}")
    print(f"  Label: {label_map_path if label_map_path else 'なし'}")
    print(f"  出力:  {out_path}")

    issues: List[str] = []

    ix_doc = parse_xml(ix_path)
    xsd_doc = parse_xml(xsd_path)
    def_doc = parse_xml(def_path)

    label_map = load_label_map_from_excel(label_map_path)
    if not label_map:
        issues.append("label_map.xlsx が無い/読めないため、日本語ラベルは空欄の可能性があります。")

    facts = collect_ix_facts(ix_doc)
    contexts = collect_contexts(ix_doc)
    units = collect_units(ix_doc)
    schema_refs = collect_schema_refs_from_ix(ix_doc)
    xsd_elements, xsd_linkbase_refs = collect_xsd_metadata(xsd_doc)
    def_locators, def_arcs = collect_def_data(def_doc)
    fact_enriched = build_fact_enriched(facts, contexts, units, label_map)
    labels_sheet = build_labels_sheet(label_map)
    concepts_without_facts = build_concepts_without_facts(xsd_elements, facts, label_map)
    hierarchy_sheet = build_hierarchy_sheet(facts, def_arcs, label_map)

    sheets = {
        "facts_all": facts,
        "facts_enriched": fact_enriched,
        "hierarchy": hierarchy_sheet,  # 新規追加：階層表示シート
        "contexts": contexts,
        "units": units,
        "schema_refs": schema_refs,
        "xsd_elements": xsd_elements,
        "xsd_linkbase_refs": xsd_linkbase_refs,
        "def_locators": def_locators,
        "def_arcs": def_arcs,
        "labels": labels_sheet,
        "concepts_without_facts": concepts_without_facts,
    }
    
    # Excel整形を自動適用
    try:
        import sys
        import os
        
        # excel_formatter.pyのパスを取得
        # 現在の位置から相対パスでexcel_formatter.pyを探す
        current_dir = Path(__file__).parent
        project_root = current_dir.parent.parent.parent.parent  # TDnet適時情報開示サービスまで遡る
        formatter_path = project_root / "python" / "excel_formatter.py"
        
        # 代替パスも試す
        if not formatter_path.exists():
            # 直接dev/pythonを探す
            formatter_path = project_root / "dev" / "python" / "excel_formatter.py"
        
        if formatter_path.exists():
            # 一時ファイルを作成して整形→出力ファイルに保存
            temp_path = out_path.with_name(out_path.stem + "_temp" + out_path.suffix)
            
            # まず一時ファイルを作成
            build_workbook(temp_path, sheets, issues)
            
            # excel_formatterをインポートして実行
            sys.path.insert(0, str(formatter_path.parent))
            from excel_formatter import ExcelFormatter
            
            print("Excel整形を適用中...")
            # サイレントモードで整形（標準出力を一時的に抑制）
            import io
            import contextlib
            
            with contextlib.redirect_stdout(io.StringIO()):
                formatter = ExcelFormatter()
                formatter.format_excel(temp_path, out_path, all_sheets=True)
            
            # 一時ファイルを削除
            temp_path.unlink()
            
            print(f"created: {out_path}")
            
        else:
            print(f"警告: Excel整形スクリプトが見つかりません: {formatter_path}")
            # 整形スクリプトがない場合は通常通り出力
            build_workbook(out_path, sheets, issues)
            print(f"created: {out_path}")
            
    except Exception as e:
        print(f"警告: Excel整形の適用に失敗しました: {e}")
        # エラーの場合は通常通り出力
        build_workbook(out_path, sheets, issues)
        print(f"created: {out_path}")
    
    return 0


if __name__ == "__main__":
    main()
