# --- obsidian_property ---
# scr名: 【自動】
# 概要: laoder.pyから売上,利益の文言を統一させCSVとExcelに出力する
# 処理grp: -
# 処理順番: 0
# mermaid: 
# tags: ["tdnet", "download"]
# aliases: ["normalizer.py"]
# created: 2026-03-31
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：laoder.pyから売上,利益の文言を統一させCSVとExcelに出力する
# --- 概要 ---


from __future__ import annotations

from pathlib import Path
from typing import Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

from loader import load_mapping, get_metric

# 設定変数
FACTS_RAW_CSV_FILE = "facts_raw.csv"    # 入力CSVファイル名
MAPPING_EXCEL_FILE = "mapping.xlsx"     # マッピングExcelファイル名
NORMALIZED_CSV_FILE = "normalized_facts.csv"    # 出力CSVファイル名
NORMALIZED_EXCEL_FILE = "normalized_facts.xlsx" # 出力Excelファイル名


def parse_context_ref(context_ref: str) -> dict[str, str | None]:
    """
    contextRef 文字列から
    - period_scope
    - consolidation
    - result_type
    を判定する。

    例:
    NextYearDuration_ConsolidatedMember_ForecastMember
    """
    text = str(context_ref or "")

    period_scope = None
    if "NextAccumulatedQ2Duration" in text:
        period_scope = "next_accumulated_q2"
    elif "NextYearDuration" in text:
        period_scope = "next_year"
    elif "CurrentAccumulatedQ2Duration" in text:
        period_scope = "current_accumulated_q2"
    elif "CurrentYearDuration" in text:
        period_scope = "current_year"
    elif "PriorYearDuration" in text:
        period_scope = "prior_year"
    elif "PriorYearAccumulatedQ2Duration" in text:
        period_scope = "prior_accumulated_q2"

    consolidation = None
    if "NonConsolidatedMember" in text:
        consolidation = "non_consolidated"
    elif "ConsolidatedMember" in text:
        consolidation = "consolidated"

    result_type = None
    if "ForecastMember" in text:
        result_type = "forecast"
    elif "ResultMember" in text:
        result_type = "result"

    return {
        "period_scope": period_scope,
        "consolidation": consolidation,
        "result_type": result_type,
    }


def normalize_value(row: pd.Series) -> Any:
    """
    facts_raw の value / scale / sign を使って数値をできるだけ正規化する。
    失敗したら元の value を返す。
    """
    value = row.get("value")
    scale = row.get("scale")
    sign = row.get("sign")

    if pd.isna(value):
        return None

    text = str(value).strip()
    if text == "":
        return None

    try:
        num = float(text.replace(",", ""))
    except ValueError:
        # 数値化できない場合は文字列のまま返す
        return value

    if pd.notna(scale) and str(scale).strip() != "":
        try:
            num *= 10 ** int(float(scale))
        except ValueError:
            pass

    if str(sign).strip() == "-":
        num *= -1

    # 整数なら int に寄せる
    if float(num).is_integer():
        return int(num)

    return num


def normalize_facts(
    facts_csv_path: str | Path,
    mapping_excel_path: str | Path,
) -> pd.DataFrame:
    """
    facts_raw.csv を読み、mapping.xlsx を使って normalized DataFrame を返す。
    """
    facts_csv_path = Path(facts_csv_path)
    mapping_excel_path = Path(mapping_excel_path)

    if not facts_csv_path.exists():
        raise FileNotFoundError(f"facts_raw.csv が見つかりません: {facts_csv_path}")

    mapping = load_mapping(mapping_excel_path)
    df = pd.read_csv(facts_csv_path, encoding="utf-8-sig")

    required_cols = {"name", "contextRef", "value"}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        raise ValueError(
            f"facts_raw.csv に必要列がありません: {sorted(missing_cols)}"
        )

    rows: list[dict[str, Any]] = []

    for _, row in df.iterrows():
        concept_name_raw = str(row.get("name", "")).strip()
        concept_name = concept_name_raw.split(":")[-1]
        metric = get_metric(concept_name, mapping)

        # マッピング対象外はスキップ
        if metric is None:
            continue

        context_ref = row.get("contextRef", "")
        ctx = parse_context_ref(str(context_ref))
        value_normalized = normalize_value(row)

        out = {
            "source_file": row.get("source_file"),
            "doc_type": row.get("doc_type"),
            "code": row.get("code"),
            "fact_type": row.get("fact_type"),
            "source_concept_raw": concept_name_raw,
            "source_concept": concept_name,
            "metric": metric,
            "contextRef": context_ref,
            "period_scope": ctx["period_scope"],
            "consolidation": ctx["consolidation"],
            "result_type": ctx["result_type"],
            "unitRef": row.get("unitRef"),
            "decimals": row.get("decimals"),
            "scale": row.get("scale"),
            "sign": row.get("sign"),
            "format": row.get("format"),
            "value_raw": row.get("value"),
            "value_normalized": value_normalized,
            "xsi_nil": row.get("xsi_nil"),
        }
        rows.append(out)

    result_df = pd.DataFrame(rows)

    if not result_df.empty:
        sort_cols = [
            col for col in
            ["code", "metric", "period_scope", "consolidation", "result_type"]
            if col in result_df.columns
        ]
        if sort_cols:
            result_df = result_df.sort_values(sort_cols).reset_index(drop=True)

    return result_df


def format_excel_sheet(sheet) -> None:
    """Excelシートを整形する
    
    Args:
        sheet: 整形対象のシート
    """
    sheet_name = sheet.title
    print(f"[INFO] シート「{sheet_name}」を整形中...")
    
    # 定数
    DEFAULT_FONT_NAME = "源ノ角ゴシック Code JP R"
    FALLBACK_FONT_NAME = "Yu Gothic"
    DEFAULT_FONT_SIZE = 10
    DEFAULT_CELL_COLOR = "FFFFFF"
    MAX_COLUMN_WIDTH = 50
    COLUMN_WIDTH_PADDING = 2
    
    # フォントと背景色の設定
    try:
        font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
        header_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    except Exception:
        font = Font(name=FALLBACK_FONT_NAME, size=DEFAULT_FONT_SIZE)
        header_font = Font(name=FALLBACK_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    
    white_fill = PatternFill("solid", fgColor=DEFAULT_CELL_COLOR)
    
    # セルの整形
    for row in sheet.iter_rows():
        for cell in row:
            # 背景を白に設定
            cell.fill = white_fill
            
            # フォント設定
            if cell.row == 1:
                # ヘッダー行は太字
                cell.font = header_font
            else:
                # データ行は通常フォント
                cell.font = font
    
    # 列幅のオートフィット
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # 列内の最大文字長を取得
        for cell in column:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        
        # 列幅を設定（余裕を持たせ、最大幅を制限）
        adjusted_width = min(max_length + COLUMN_WIDTH_PADDING, MAX_COLUMN_WIDTH)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # グリッド線を非表示
    sheet.sheet_view.showGridLines = False
    
    print(f"[OK] シート「{sheet_name}」の整形完了")


def main() -> None:
    base_dir = Path(__file__).parent

    facts_csv_path = base_dir / FACTS_RAW_CSV_FILE
    mapping_excel_path = base_dir / MAPPING_EXCEL_FILE
    output_csv_path = base_dir / NORMALIZED_CSV_FILE
    output_excel_path = base_dir / NORMALIZED_EXCEL_FILE

    df_normalized = normalize_facts(
        facts_csv_path=facts_csv_path,
        mapping_excel_path=mapping_excel_path,
    )

    # CSV出力
    df_normalized.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
    
    # Excel出力
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_normalized.to_excel(writer, sheet_name='正規化データ', index=False)
        
        # 来期予想データのみのシートも作成
        forecast_df = df_normalized[
            (df_normalized["period_scope"] == "next_year")
            & (df_normalized["result_type"] == "forecast")
        ]
        if not forecast_df.empty:
            forecast_df.to_excel(writer, sheet_name='来期予想', index=False)
            
            # 1行にまとめた来期予想シートを作成
            forecast_pivot = forecast_df.pivot_table(
                index=['code', 'consolidation'],
                columns='metric',
                values='value_normalized',
                aggfunc='first'
            ).reset_index()
            
            # source_file列を追加
            forecast_pivot['source_file'] = forecast_pivot['code'].map(
                forecast_df.drop_duplicates('code').set_index('code')['source_file']
            )
            
            # 重複文言列を追加（同じsource_file内で連続した行に同じcodeがある場合のみ）
            forecast_pivot['重複'] = ''
            
            # source_fileでソートして連続した重複をチェック
            forecast_sorted = forecast_pivot.sort_values('source_file').reset_index(drop=True)
            
            for i in range(len(forecast_sorted) - 1):
                current_row = forecast_sorted.iloc[i]
                next_row = forecast_sorted.iloc[i + 1]
                
                # 同じsource_fileで同じcodeが連続している場合
                if (current_row['source_file'] == next_row['source_file'] and 
                    current_row['code'] == next_row['code']):
                    
                    # 該当行のインデックスを取得して重複マーク
                    current_idx = forecast_pivot[forecast_pivot['code'] == current_row['code']].index
                    next_idx = forecast_pivot[forecast_pivot['code'] == next_row['code']].index
                    
                    forecast_pivot.loc[current_idx, '重複'] = '重複'
                    forecast_pivot.loc[next_idx, '重複'] = '重複'
            
            # 指標の順番を整える
            metric_order = ['売上', '営業利益', '経常利益', '純利益', '純利益親会社株主']
            available_metrics = [m for m in metric_order if m in forecast_pivot.columns]
            final_columns = ['source_file', '重複', 'code', 'consolidation'] + available_metrics
            forecast_pivot = forecast_pivot[final_columns]
            
            forecast_pivot.to_excel(writer, sheet_name='来期予想_1行', index=False)

    # Excel整形処理
    print("\n=== Excel整形開始 ===")
    wb = openpyxl.load_workbook(output_excel_path)
    for sheet in wb.worksheets:
        format_excel_sheet(sheet)
    wb.save(output_excel_path)
    print("=== Excel整形完了 ===")

    print("=== 正規化完了 ===")
    print(f"入力: {facts_csv_path}")
    print(f"CSV出力: {output_csv_path}")
    print(f"Excel出力: {output_excel_path}")
    print(f"件数: {len(df_normalized)}")

    if not df_normalized.empty:
        print()
        print("=== 先頭10件 ===")
        print(
            df_normalized[
                [
                    "code",
                    "source_concept_raw",
                    "source_concept",
                    "metric",
                    "period_scope",
                    "consolidation",
                    "result_type",
                    "value_raw",
                    "value_normalized",
                ]
            ].head(10).to_string(index=False)
        )

        print()
        print("=== 来期予想だけ確認 ===")
        forecast_df = df_normalized[
            (df_normalized["period_scope"] == "next_year")
            & (df_normalized["result_type"] == "forecast")
        ]

        if forecast_df.empty:
            print("next_year + forecast のデータは見つかりませんでした。")
        else:
            cols = [
                "code",
                "source_concept_raw",
                "source_concept",
                "metric",
                "consolidation",
                "value_normalized",
            ]
            print(forecast_df[cols].head(20).to_string(index=False))


if __name__ == "__main__":
    main()