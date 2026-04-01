# --- obsidian_property ---
# scr名: 【自動】
# 概要: 全タグと属性を抽出
# 処理grp: XBRL
# 処理順番: -
# input: XML, XSD, iXBRL, def.xmlファイル指定
# output: excelファイル
# mermaid: 
# tags: ["XBRL", "python"]
# aliases: 
# created: 2026-03-23
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：全タグと属性を抽出
# 1sheet目: 対象ファイル
# 2sheet目: タグ/属性リスト
# 3sheet目以降: 各タグの詳細情報
# --- 概要 ---

from __future__ import annotations

import io
import re
import subprocess
import sys
import urllib.request
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List


# ======================================================
# 設定: 処理対象ファイル／URL と出力ファイルをここで指定
# ======================================================
# フォルダ設定（変数化）
BASE_FOLDER = r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス"  # 基準フォルダパス
INPUT_FOLDER = r"決算短信"  # 入力ファイル格納フォルダ（BASE_FOLDERからの相対パス）
OUTPUT_FOLDER = Path(__file__).parent  # スクリプトと同じ場所

# ファイル設定（変数化）
SOURCES = [
    # ローカルファイルパス例:
    # r"tse-acedjpsm-71720-20260209552878.xsd",
    # r"tse-acedjpsm-71720-20260209552878-def.xml",
    # r"tse-acedjpsm-71720-20260209552878-ixbrl.htm"
    
    # URL例:
    "http://www.xbrl.tdnet.info/taxonomy/jp/tse/tdnet/ed/t/2014-01-12/tse-ed-t-2014-01-12.xsd",
    #"http://www.xbrl.tdnet.info/taxonomy/jp/tse/tdnet/ed/o/rt/2014-01-12/tse-ed-rt-2014-01-12.xsd",
    #"http://www.xbrl.tdnet.info/taxonomy/jp/tse/tdnet/ed/t/2014-01-12/tse-ed-t-2014-01-12-lab.xml",
]
def get_output_filename():
    """日時付きの出力ファイル名を生成"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = Path(SOURCES[0]).name
    return f"tag_data_get.py_{base_name}_{timestamp}.xlsx"

OUTPUT_FILE = get_output_filename()
ENCODING = "utf-8"

# Excel整形設定
USE_FORMATTER = True  # Trueの場合、excel_formatter.pyで整形する
# ======================================================

# 変数を簡単に変更できるように関数化
def get_settings():
    """設定を返す関数"""
    return {
        'base_folder': BASE_FOLDER,
        'input_folder': INPUT_FOLDER,
        'output_folder': OUTPUT_FOLDER,
        'sources': SOURCES,
        'output_file': OUTPUT_FILE,
        'encoding': ENCODING,
        'use_formatter': USE_FORMATTER
    }

def format_excel_with_formatter(input_path: Path, output_path: Path) -> bool:
    """整形スクリプトを呼び出してExcelを整形する"""
    try:
        settings = get_settings()
        formatter_script = Path(settings['base_folder']).parent / "python" / "excel_formatter.py"
        cmd = [
            sys.executable, 
            str(formatter_script),
            "--input", str(input_path),
            "--output", str(output_path),
            "--all-sheets"
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='cp932', errors='replace')
        
        if result.returncode == 0:
            print(f"[OK] Excelを整形しました: {output_path}")
            return True
        else:
            print(f"[ERROR] 整形に失敗しました: {result.stderr}")
            return False
            
    except Exception as exc:
        print(f"[ERROR] 整形スクリプト実行エラー: {exc}")
        return False

from openpyxl import Workbook
from openpyxl.styles import Font


# よく使うnamespaceを接頭辞つき属性名に戻すための対応表
NS_PREFIX_MAP = {
    "http://www.w3.org/1999/xlink": "xlink",
    "http://www.xbrl.org/2003/instance": "xbrli",
    "http://xbrl.org/2005/xbrldt": "xbrldt",
    "http://www.w3.org/2001/XMLSchema": "xs",
    "http://www.w3.org/2001/XMLSchema-instance": "xsi",
    "http://www.xbrl.org/2003/linkbase": "link",
    "http://www.xbrl.org/2008/inlineXBRL": "ix",
    "http://www.xbrl.org/inlineXBRL/transformation/2011-07-31": "ixt",
    "http://www.xbrl.org/dtr/type/numeric": "num",
    "http://www.xbrl.org/dtr/type/non-numeric": "nonnum",
}

# よく使う属性は左側に寄せる
PREFERRED_ATTR_ORDER = [
    "name",
    "id",
    "type",
    "substitutionGroup",
    "abstract",
    "nillable",
    "xlink:type",
    "xlink:href",
    "xlink:label",
    "xlink:from",
    "xlink:to",
    "xlink:arcrole",
    "xlink:role",
    "contextRef",
    "unitRef",
    "decimals",
    "scale",
    "format",
    "order",
    "priority",
    "use",
    "xbrli:balance",
    "xbrli:periodType",
    "xbrldt:closed",
    "xbrldt:contextElement",
]


def is_url(path: str) -> bool:
    return path.startswith("http://") or path.startswith("https://")


def read_text_from_source(source: str, encoding: str = "utf-8") -> str:
    if is_url(source):
        with urllib.request.urlopen(source) as resp:
            raw = resp.read()
        return raw.decode(encoding, errors="replace")
    
    # ローカルファイルの場合、フォルダ設定を適用
    settings = get_settings()
    base_path = Path(settings['base_folder'])
    input_path = base_path / settings['input_folder'] / source
    return input_path.read_text(encoding=encoding, errors="replace")


def strip_ns(tag: str) -> str:
    """{namespace}local -> local"""
    if tag.startswith("{"):
        return tag.rsplit("}", 1)[1]
    return tag


def normalize_attr_name(attr_name: str) -> str:
    """
    ElementTreeの属性名:
      '{namespace}href' -> 'xlink:href'
      'name' -> 'name'
    に戻す
    """
    if attr_name.startswith("{"):
        ns, local = attr_name[1:].split("}", 1)
        prefix = NS_PREFIX_MAP.get(ns)
        return f"{prefix}:{local}" if prefix else f"{{{ns}}}{local}"
    return attr_name


def attrs_to_normalized_dict(elem: ET.Element) -> Dict[str, str]:
    return {normalize_attr_name(k): v for k, v in elem.attrib.items()}


def parse_all_elements_from_text(
    xml_text: str,
    source_name: str,
) -> Dict[str, List[Dict[str, str]]]:
    """
    全タグを抽出して、タグごとの行データを返す。
    """
    results: Dict[str, List[Dict[str, str]]] = defaultdict(list)

    stream = io.StringIO(xml_text)
    context = ET.iterparse(stream, events=("start",))

    for _event, elem in context:
        local_tag = strip_ns(elem.tag)

        row = {
            "_source": source_name,
            "_tag": local_tag,
        }
        row.update(attrs_to_normalized_dict(elem))

        text = (elem.text or "").strip()
        if text:
            row["_text"] = text

        results[local_tag].append(row)

    return results


def merge_results(
    all_results: Iterable[Dict[str, List[Dict[str, str]]]]
) -> Dict[str, List[Dict[str, str]]]:
    merged: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for result in all_results:
        for tag, rows in result.items():
            merged[tag].extend(rows)
    return merged


def sort_headers(headers: Iterable[str]) -> List[str]:
    headers = list(headers)

    fixed_front = ["_source", "_tag", "_text"]
    present_front = [h for h in fixed_front if h in headers]

    preferred = [h for h in PREFERRED_ATTR_ORDER if h in headers]
    remaining = sorted(
        [h for h in headers if h not in set(present_front) | set(preferred)]
    )

    return present_front + preferred + remaining


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    r"""
    Excelシート名制約:
    - 31文字以内
    - : \ / ? * [ ] を含めない
    - 重複不可
    """
    cleaned = re.sub(r'[:\\/*?\[\]]', "_", name).strip()
    if not cleaned:
        cleaned = "blank"

    cleaned = cleaned[:31]

    if cleaned not in used_names:
        used_names.add(cleaned)
        return cleaned

    i = 2
    while True:
        suffix = f"_{i}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        i += 1


def autofit_rough(ws) -> None:
    """
    openpyxlでは厳密な自動幅調整がないので、ざっくり幅を設定
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def create_statistics_data(merged: Dict[str, List[Dict[str, str]]]) -> List[List]:
    """統計データを作成する"""
    stats_data = []
    
    for tag in sorted(merged.keys()):
        rows = merged[tag]
        if not rows:
            continue
            
        # タグの出現回数
        tag_count = len(rows)
        
        # 属性の出現回数を集計
        attr_counts = defaultdict(int)
        for row in rows:
            for attr_name in row.keys():
                if not attr_name.startswith('_'):  # _source, _tag, _text は除外
                    if row[attr_name]:  # 空でない属性のみカウント
                        attr_counts[attr_name] += 1
        
        # タグの1行目：タグ出現回数
        stats_data.append([tag, "", tag_count])
        
        # 各属性の出現回数
        for attr_name in sorted(attr_counts.keys()):
            stats_data.append([tag, attr_name, attr_counts[attr_name]])
    
    return stats_data


def write_excel_per_tag(
    output_xlsx: Path,
    merged: Dict[str, List[Dict[str, str]]],
    sources: List[str],
) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    used_sheet_names: set[str] = set()

    # 1番目のシート：参照元シート
    source_ws = wb.create_sheet(title="参照元", index=0)
    source_ws.freeze_panes = "A2"
    
    # 参照元シートのヘッダー
    source_ws.append(["参照元"])
    
    # 参照元データを追加
    for source in sources:
        source_ws.append([source])
    
    autofit_rough(source_ws)
    used_sheet_names.add("参照元")

    # 2番目のシート：統計シート
    stats_ws = wb.create_sheet(title="統計", index=1)
    stats_ws.freeze_panes = "A2"
    
    # 統計シートのヘッダー
    stats_ws.append(["タグ名", "属性名", "出現回数"])
    
    # 統計データを作成して追加
    stats_data = create_statistics_data(merged)
    tag_count = 0  # タグ数をカウント
    current_row = 2  # ヘッダーの次の行から開始
    
    for row_data in stats_data:
        stats_ws.append(row_data)

        # B列がブランク（タグ名のトップ）の行を太字にする
        if row_data[1] == "":  # B列（属性名）がブランクの場合
            for col in range(1, 4):  # A, B, C列を太字にする
                cell = stats_ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True)
            tag_count += 1
        
        current_row += 1
    
    # 最終行にタグ数の合計を追加
    final_row = current_row  # 空白行なし
    stats_ws.cell(row=final_row, column=1, value="合計")
    # SUMIF関数でB列がブランクのC列を合計
    stats_ws.cell(row=final_row, column=3, value=f'=SUMIF(B2:B{current_row-1},"",C2:C{current_row-1})')
    
    # 合計行も太字にする
    for col in range(1, 4):
        cell = stats_ws.cell(row=final_row, column=col)
        cell.font = Font(bold=True)
    
    autofit_rough(stats_ws)
    used_sheet_names.add("統計")

    # 各タグごとのシート
    for tag in sorted(merged.keys()):
        rows = merged[tag]
        if not rows:
            continue

        all_headers = set()
        for row in rows:
            all_headers.update(row.keys())

        headers = sort_headers(all_headers)

        ws = wb.create_sheet(title=safe_sheet_name(tag, used_sheet_names))
        ws.freeze_panes = "A2"

        # header
        ws.append(headers)

        # rows
        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        autofit_rough(ws)

    # 一時ファイルに保存
    temp_output_path = output_xlsx.with_suffix('.temp.xlsx')
    temp_output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(temp_output_path)
    
    # 整形スクリプトを呼び出して最終ファイルを作成
    settings = get_settings()
    if settings['use_formatter']:
        if format_excel_with_formatter(temp_output_path, output_xlsx):
            # 一時ファイルを削除
            temp_output_path.unlink()
            print(f"[OK] Excelを整形して保存しました: {output_xlsx}")
        else:
            print(f"[ERROR] 整形に失敗したため、一時ファイルを保持: {temp_output_path}")
    else:
        # 整形しない場合は一時ファイルをリネーム
        temp_output_path.replace(output_xlsx)
        print(f"[OK] Excel saved -> {output_xlsx}")


def main() -> int:
    # 設定を取得
    settings = get_settings()
    sources = settings['sources']
    
    if not sources:
        print("SOURCES が空です。スクリプト上部の SOURCES にファイルパスまたはURLを指定してください。", file=sys.stderr)
        return 1

    # 出力フォルダの作成
    base_path = Path(settings['base_folder'])
    output_folder = settings['output_folder']
    
    if output_folder == "":
        # スクリプトと同じ場所に出力
        output_path = Path(__file__).parent
    else:
        # 指定されたフォルダに出力
        output_path = base_path / output_folder
    
    output_path.mkdir(parents=True, exist_ok=True)
    
    # 出力ファイルのフルパスを設定
    output_file_path = output_path / settings['output_file']

    all_results = []

    for source in sources:
        try:
            xml_text = read_text_from_source(source, encoding=settings['encoding'])
            result = parse_all_elements_from_text(
                xml_text=xml_text,
                source_name=source,
            )
            all_results.append(result)
            found = sum(len(v) for v in result.values())
            print(f"[READ] {source} -> {found} elements")
        except Exception as e:
            print(f"[ERROR] {source}: {e}", file=sys.stderr)

    merged = merge_results(all_results)

    if not merged:
        print("要素は見つかりませんでした。", file=sys.stderr)
        return 2

    write_excel_per_tag(output_file_path, merged, sources)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())