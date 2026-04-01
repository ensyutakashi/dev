# --- obsidian_property ---
# scr名: 【自動】
# 概要: TDnetの-ixbrl.htmからcontextを抽出してタグとValueを抽出
# 処理grp: XBRL決算短信取得
# 処理順番: -
# mermaid: 
# tags: ["tdnet", "決算短信", "tag管理"]
# aliases: ["tdnet_get_xblr_context.py"]
# created: 2026-03-26
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：TDnetの-ixbrl.htmからcontentを抽出してタグとValueを抽出
#　XBRLのcontextにどのようなタグとvalueがあるかを全リスト化する
#　漏れ防止、これをもとにタグとvalueのマッピングを作成する
# --- 概要 ---

from datetime import datetime
from lxml import etree
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# 設定変数
LIMIT = 50000  # 処理するファイル数の上限

# 入力フォルダ
base_folder = str(Path(__file__).parent) #このスクリプトと同じフォルダ
# 入力ファイル
input_file_path = str(Path(__file__).parent / "tse-acedjpsm-71720-20260209552878-ixbrl.htm")

#出力
output_folder_path = str(Path(__file__).parent)
output_excel_filename = f"{Path(__file__).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# 名前空間
ns = {
    "xbrli": "http://www.xbrl.org/2003/instance",
    "xbrldi": "http://xbrl.org/2006/xbrldi",
}

# 全てのコンテキスト情報を格納するリスト
all_rows = []

# 処理したファイル数をカウント
file_count = 0

# 単一ファイル処理か一括処理かを判定
if input_file_path:
    target_path = Path(input_file_path)
    if not target_path.exists():
        raise FileNotFoundError(f"指定されたファイルが見つかりません: {target_path}")
    ixbrl_files = [target_path]
    print(f"単一ファイルを処理します: {target_path.name}")
else:
    # ベースフォルダのPathオブジェクト
    base_path = Path(base_folder)
    
    # _決算短信_を含むフォルダを検索（サブフォルダも含む）
    kessan_folders = [d for d in base_path.rglob("*") if d.is_dir() and "_決算短信_" in d.name]
    
    print(f"検索対象フォルダ数: {len(kessan_folders)}")
    
    ixbrl_files = []
    for folder in kessan_folders:
        # Summaryフォルダを探す
        summary_folder = folder / "XBRLData" / "Summary"
        
        if not summary_folder.exists():
            continue
        
        # Summary内のixbrl.htmファイルを探す
        ixbrl_files.extend(list(summary_folder.glob("*-ixbrl.htm")))

# 各ファイルを処理
for ixbrl_file in ixbrl_files:
    if file_count >= LIMIT:
        break
    
    try:
        print(f"\n処理中 ({file_count + 1}/{LIMIT}): {ixbrl_file.name}")
        
        # XMLをパース
        tree = etree.parse(str(ixbrl_file))
        root = tree.getroot()
        
        # コンテキストタグを抽出
        contexts = root.xpath(".//xbrli:context", namespaces=ns)
        
        print(f"  コンテキスト数: {len(contexts)}")
        
        for ctx in contexts:
            context_id = ctx.get("id", "")
            
            identifier = ctx.findtext(".//xbrli:identifier", namespaces=ns, default="")
            
            start_date = ctx.findtext(".//xbrli:startDate", namespaces=ns, default="")
            end_date   = ctx.findtext(".//xbrli:endDate", namespaces=ns, default="")
            instant    = ctx.findtext(".//xbrli:instant", namespaces=ns, default="")
            
            # dimensionを辞書として解析
            dimensions = {}
            for m in ctx.xpath(".//xbrldi:explicitMember", namespaces=ns):
                dim = m.get("dimension", "")
                member = (m.text or "").strip()
                
                # dimension名からプレフィックスを除去して列名にする
                if ":" in dim:
                    dim_name = dim.split(":")[-1]
                else:
                    dim_name = dim
                    
                # member値からプレフィックスを除去
                if ":" in member:
                    member_value = member.split(":")[-1]
                else:
                    member_value = member
                    
                dimensions[dim_name] = member_value
            
            # 基本データとdimensionsをマージ
            row_data = {
                "file_name": ixbrl_file.name,
                "folder_name": folder.name if not input_file_path else Path(input_file_path).parent.name,
                "context_id": context_id,
                "identifier": identifier,
                "start_date": start_date,
                "end_date": end_date,
                "instant": instant,
            }
            
            # dimensionデータを追加
            row_data.update(dimensions)
            
            all_rows.append(row_data)
        
        file_count += 1
        
    except Exception as e:
        print(f"  エラー: {e}")
        continue

print(f"\n\n処理完了: {file_count}ファイル処理")
print(f"抽出したコンテキスト総数: {len(all_rows)}")

# DataFrameを作成
df = pd.DataFrame(all_rows)
print(f"\nDataFrame作成完了: {len(df)}行")

# 出力フォルダを作成
output_folder = Path(output_folder_path)
output_folder.mkdir(parents=True, exist_ok=True)

# 一時Excelファイルに出力
temp_excel_filename = f"temp_{output_excel_filename}"
temp_excel_path = output_folder / temp_excel_filename
final_excel_path = output_folder / output_excel_filename

df.to_excel(temp_excel_path, index=False, engine='openpyxl')
print(f"\n一時Excelファイルに出力しました: {temp_excel_path}")

# Excel整形処理
try:
    print("\nExcel整形処理を開始...")
    
    # 入力ファイルを読み込み
    wb = load_workbook(temp_excel_path)
    
    # フォント設定（源ノ角ゴシック Code JP R）
    try:
        font_name = "源ノ角ゴシック Code JP R"
        font = Font(name=font_name, size=10)
    except:
        font_name = "Yu Gothic"
        font = Font(name=font_name, size=10)
    
    # 背景を白に設定
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    
    # アクティブシートを整形
    sheet = wb.active
    print(f"[INFO] シート「{sheet.title}」を整形中...")
    
    # 使用済みセルの背景を白に設定
    for row in sheet.iter_rows():
        for cell in row:
            cell.fill = white_fill
            # ヘッダー行（1行目）は太字にする
            if cell.row == 1:
                cell.font = Font(name=font_name, size=10, bold=True)
            else:
                cell.font = font
    
    # 列のオートフィット
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value is not None:
                # セルの値を文字列に変換して長さを取得
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # 列幅を設定（少し余裕を持たせる）
        adjusted_width = min(max_length + 2, 50)  # 最大50文字に制限
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # グリッド線を非表示
    sheet.sheet_view.showGridLines = False
    
    print(f"[OK] シート「{sheet.title}」の整形完了")
    
    # 整形済みファイルを保存
    wb.save(final_excel_path)
    print(f"[OK] Excelを整形しました: {final_excel_path}")
    
    # 一時ファイルを削除
    temp_excel_path.unlink()
    print(f"[INFO] 一時ファイルを削除しました")
    
except Exception as exc:
    print(f"[ERROR] 整形に失敗しました: {exc}")
    print(f"[INFO] 一時ファイルを保持: {temp_excel_path}")
