# --- obsidian_property ---
# scr名: 【自動】
# 概要: TDnetの-ixbrl.htmからnonFraction/nonNumericの属性を抽出
# 処理grp: XBRL決算短信取得
# 処理順番: -
# mermaid:
# tags: ["tdnet", "決算短信", "tag管理"]
# aliases: ["tdnet_get_xbrl_attribute.py"]
# created: 2026-03-26
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：TDnetの-ixbrl.htmからnonFraction/nonNumeric要素の属性を抽出
#　contextRef, name, unitRef, format, decimals, scale, xsi:nil, escape, signが抽出された
# --- 概要 ---

from collections import Counter
from datetime import datetime
from lxml import etree
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ==============================
# 🔧 設定（ここだけ変更）
# ==============================
CONFIG = {
    "limit": 500000,  # 処理するファイル数の上限
    # 入力データ元
    "base_folder": r"C:\Users\ensyu\_myfolder\work\dev\TDnet適時情報開示サービス\00_TDnet(決算短信)XBRL",
    # 出力先
    "output_folder_path": str(Path(__file__).parent),
    "output_excel_filename": f"{Path(__file__).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    # 対象条件
    "folder_keyword": "_決算短信_",
    "summary_subdir": "Summary",
    "file_glob": "*-ixbrl.htm",
    # 単一ファイル処理の場合に指定（空なら一括処理）
    "input_file_path": "",
}

LIMIT = CONFIG["limit"]
base_folder = CONFIG["base_folder"]
output_folder_path = CONFIG["output_folder_path"]
output_excel_filename = CONFIG["output_excel_filename"]
FOLDER_KEYWORD = CONFIG["folder_keyword"]
SUMMARY_SUBDIR = CONFIG["summary_subdir"]
FILE_GLOB = CONFIG["file_glob"]
input_file_path = CONFIG["input_file_path"]

# 名前空間
ns = {
    "xhtml": "http://www.w3.org/1999/xhtml",
    "xbrli": "http://www.xbrl.org/2003/instance",
    "xbrldi": "http://xbrl.org/2006/xbrldi",
}
ix_uris = [
    "http://www.xbrl.org/2013/inlineXBRL",
    "http://www.xbrl.org/2008/inlineXBRL",
]


def clean_text(text: str) -> str:
    return " ".join(text.split()).strip()


def infer_folder_name(ixbrl_file: Path) -> str:
    # .../XBRLData/Summary/xxx-ixbrl.htm の場合は親フォルダ名を返す。
    parts = [p.name for p in ixbrl_file.parents]
    if "XBRLData" in parts:
        idx = parts.index("XBRLData")
        if idx + 1 < len(parts):
            return parts[idx + 1]
    return ixbrl_file.parent.name


ixbrl_targets = []
start_time = datetime.now()
print(f"開始時間: {start_time:%Y-%m-%d %H:%M:%S}")

if input_file_path:
    target_path = Path(input_file_path)
    if not target_path.exists():
        raise FileNotFoundError(f"指定されたファイルが見つかりません: {target_path}")
    ixbrl_targets = [target_path]
else:
    base_path = Path(base_folder)
    kessan_folders = [
        d for d in base_path.rglob("*")
        if d.is_dir() and FOLDER_KEYWORD in d.name
    ]
    print(f"検索対象フォルダ数: {len(kessan_folders)}")

    for folder in kessan_folders:
        summary_folder = folder / "XBRLData" / SUMMARY_SUBDIR
        if not summary_folder.exists():
            continue
        ixbrl_targets.extend(list(summary_folder.glob(FILE_GLOB)))


attribute_counts = Counter()
file_count = 0

for ixbrl_file in sorted(ixbrl_targets):
    if file_count >= LIMIT:
        print(f"\nリミット({LIMIT}件)に達したため処理を終了します")
        break

    try:
        print(f"\n処理中 ({file_count + 1}/{LIMIT}): {ixbrl_file.name}")
        parser = etree.XMLParser(recover=True, huge_tree=True)
        tree = etree.parse(str(ixbrl_file), parser)
        root = tree.getroot()

        facts = []
        for uri in ix_uris:
            facts.extend(
                root.xpath(
                    f"//*[namespace-uri()='{uri}' and (local-name()='nonFraction' or local-name()='nonNumeric')]"
                )
            )

        print(f"  抽出fact数: {len(facts)}")

        for elem in facts:
            for key in elem.attrib.keys():
                if key == "{http://www.w3.org/2001/XMLSchema-instance}nil":
                    attribute_counts["xsi:nil"] += 1
                else:
                    attribute_counts[key] += 1

        file_count += 1

    except Exception as e:
        print(f"  エラー: {e}")
        continue

end_time = datetime.now()
elapsed = end_time - start_time

print(f"\n\n処理完了: {file_count}ファイル処理")
print(f"抽出属性数: {len(attribute_counts)}")
print(f"終了時間: {end_time:%Y-%m-%d %H:%M:%S}")
print(f"処理時間: {elapsed}")

df = pd.DataFrame(
    sorted(attribute_counts.items()),
    columns=["attribute", "count"],
)
print(f"\nDataFrame作成完了: {len(df)}行")

output_folder = Path(output_folder_path)
output_folder.mkdir(parents=True, exist_ok=True)

temp_excel_filename = f"temp_{output_excel_filename}"
temp_excel_path = output_folder / temp_excel_filename
final_excel_path = output_folder / output_excel_filename

df.to_excel(temp_excel_path, index=False, engine='openpyxl')
print(f"\n一時Excelファイルに出力しました: {temp_excel_path}")

try:
    print("\nExcel整形処理を開始...")
    wb = load_workbook(temp_excel_path)

    try:
        font_name = "源ノ角ゴシック Code JP R"
        font = Font(name=font_name, size=10)
    except Exception:
        font_name = "Yu Gothic"
        font = Font(name=font_name, size=10)

    white_fill = PatternFill("solid", fgColor="FFFFFF")

    sheet = wb.active
    print(f"[INFO] シート「{sheet.title}」を整形中...")

    for row in sheet.iter_rows():
        for cell in row:
            cell.fill = white_fill
            if cell.row == 1:
                cell.font = Font(name=font_name, size=10, bold=True)
            else:
                cell.font = font

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width

    sheet.sheet_view.showGridLines = False

    print(f"[OK] シート「{sheet.title}」の整形完了")

    wb.save(final_excel_path)
    print(f"[OK] Excelを整形しました: {final_excel_path}")

    temp_excel_path.unlink()
    print("[INFO] 一時ファイルを削除しました")

except Exception as exc:
    print(f"[ERROR] 整形に失敗しました: {exc}")
    print(f"[INFO] 一時ファイルを保持: {temp_excel_path}")
