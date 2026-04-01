# --- obsidian_property ---
# scr名: 【自動】
# 概要: XBRL Labelling Linkbase生成
# 処理grp: XBRL
# 処理順番: 0
# input: 無し
# output: label_map.xlsx
# mermaid: 
# tags: ["XBRL", "python"]
# aliases: 
# created: 2026-03-23
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：XBRL Labelling Linkbase生成
# TDnetのXBRLラベルXMLをダウンロードし、Excelに整形して出力する
# --- 概要 ---

from __future__ import annotations

from pathlib import Path
from collections import defaultdict
import urllib.request
import xml.etree.ElementTree as ET
import subprocess
import sys
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LAB_XML_URL = "http://www.xbrl.tdnet.info/taxonomy/jp/tse/tdnet/ed/t/2014-01-12/tse-ed-t-2014-01-12-lab.xml"

# 出力設定（デフォルトはスクリプトと同じ場所）
OUTPUT_DIR = Path(__file__).parent  # スクリプトと同じ場所
OUTPUT_FILENAME = "label_map.xlsx"
OUTPUT_FILE = OUTPUT_DIR / OUTPUT_FILENAME

NS = {
    "link": "http://www.xbrl.org/2003/linkbase",
    "xlink": "http://www.w3.org/1999/xlink",
    "xml": "http://www.w3.org/XML/1998/namespace",
}


def download_xml(url: str, timeout: int = 30) -> bytes:
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0"}
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read()


def concept_from_href(href: str) -> str:
    """
    例:
      http://...#tse-ed-t_CompanyName -> CompanyName
    """
    raw = href.split("#")[-1]
    return raw.split("_", 1)[1] if "_" in raw else raw


def parse_lab_xml(xml_bytes: bytes) -> list[dict]:
    root = ET.fromstring(xml_bytes)

    rows: list[dict] = []

    # labelLinkごとに loc / label / labelArc を対応付ける
    for label_link in root.findall(".//link:labelLink", NS):
        loc_map: dict[str, str] = {}
        label_res_map: dict[str, dict] = {}

        for loc in label_link.findall("link:loc", NS):
            loc_label = loc.get(f"{{{NS['xlink']}}}label")
            href = loc.get(f"{{{NS['xlink']}}}href", "")
            if loc_label:
                loc_map[loc_label] = concept_from_href(href)

        for label in label_link.findall("link:label", NS):
            res_label = label.get(f"{{{NS['xlink']}}}label")
            label_role = label.get(f"{{{NS['xlink']}}}role", "")
            xml_lang = label.get(f"{{{NS['xml']}}}lang", "")
            text = (label.text or "").strip()

            if res_label:
                label_res_map[res_label] = {
                    "label_text": text,
                    "label_role": label_role,
                    "xml_lang": xml_lang,
                }

        for arc in label_link.findall("link:labelArc", NS):
            frm = arc.get(f"{{{NS['xlink']}}}from")
            to = arc.get(f"{{{NS['xlink']}}}to")

            concept = loc_map.get(frm or "", "")
            label_info = label_res_map.get(to or "", {})

            if concept and label_info.get("label_text"):
                rows.append({
                    "concept": concept,
                    "日本語ラベル": label_info["label_text"],
                    "label_role": label_info["label_role"],
                    "xml_lang": label_info["xml_lang"],
                })

    # 重複排除
    seen = set()
    unique_rows = []
    for row in rows:
        key = (
            row["concept"],
            row["日本語ラベル"],
            row["label_role"],
            row["xml_lang"],
        )
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    # 標準ラベル優先 + concept順
    def sort_key(r: dict):
        role = r["label_role"]
        standard_first = 0 if role.endswith("/label") else 1
        return (r["concept"], standard_first, role, r["日本語ラベル"])

    unique_rows.sort(key=sort_key)
    return unique_rows


def autofit_columns(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def write_excel(rows: list[dict], output_file: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "label_map"

    headers = ["concept", "日本語ラベル", "label_role", "xml_lang"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["concept"],
            row["日本語ラベル"],
            row["label_role"],
            row["xml_lang"],
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    autofit_columns(ws)

    wb.save(output_file)


def format_excel_with_formatter(input_path: Path, output_path: Path) -> bool:
    """整形スクリプトを呼び出してExcelを整形する"""
    try:
        formatter_script = Path(__file__).parent.parent.parent.parent / "python" / "excel_formatter.py"
        cmd = [
            sys.executable, 
            str(formatter_script),
            "--input", str(input_path),
            "--output", str(output_path),
            "--all-sheets"
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='cp932', errors='replace')
        
        if result.returncode == 0:
            print(f"[OK] Excelを整形しました: {output_path}")
            return True
        else:
            print(f"[ERROR] 整形に失敗しました: {result.stderr}")
            return False
            
    except Exception as exc:
        print(f"[ERROR] 整形スクリプト実行エラー: {exc}")
        return False


def main():
    xml_bytes = download_xml(LAB_XML_URL)
    rows = parse_lab_xml(xml_bytes)
    
    # 出力フォルダを作成
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # 一時ファイルに出力
    temp_file = OUTPUT_FILE.with_suffix(".temp.xlsx")
    write_excel(rows, temp_file)
    print(f"一時出力完了: {temp_file.resolve()}")
    print(f"件数: {len(rows)}")
    
    # excel_formatter.py で整形
    if format_excel_with_formatter(temp_file, OUTPUT_FILE):
        # 一時ファイルを削除
        temp_file.unlink()
        print("一時ファイルを削除しました")
    else:
        print(f"整形に失敗したため、一時ファイルを残します: {temp_file.resolve()}")

if __name__ == "__main__":
    main()