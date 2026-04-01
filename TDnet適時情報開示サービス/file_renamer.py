# --- obsidian_property ---
# scr名: 【自動】
# 概要: Excelファイルを使ってファイル名を一括変更するツール
# 処理grp: tool
# 処理順番: 0
# mermaid: -
# tags: ["rename"]
# aliases: ["file_renamer.py"]
# created: 2026-02-26
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：
# Excelファイルを使ってファイル名を一括変更するツール
# --- 概要 ---

import os
import openpyxl
from pathlib import Path
import sys

# ==========================================
# 設定変数 (ここを変更して実行してください)
# ==========================================
# 1. エクセルファイルが置いてあるフォルダパス
# デフォルト（Path(__file__).parent）はスクリプトと同じ場所です
EXCEL_DIR = Path(__file__).parent 

# 2. エクセルファイルのファイル名
# EXCEL_NAME = "file_renamePDF_20260225.xlsx"
EXCEL_NAME = "file_renamer_format.xlsx"

# 3. 対象ファイルの拡張子（ドットを含めて指定）
FILE_EXTENSION = ".zip"

# 4. 対象フラグの名称（D列）
TARGET_LABEL = "変更対象"
SKIP_LABEL   = "対象外"

# 5. エクセルのシート名（Noneの場合は一番左のアクティブなシート）
SHEET_NAME = None

# ==========================================
# クラス定義
# ==========================================

class ExcelFileRenamer:
    """
    Excelファイルを使ってファイル名を一括変更するツール
    A列: 対象のフォルダパス
    B列: 変更前名称
    C列: 変更後名称
    D列: 変更対象の有無
    E列: 処理結果
    """
    
    def __init__(self, excel_path=None, sheet_name=None, target_label="変更対象", 
                 skip_label="対象外", file_extension=".pdf"):
        self.excel_path = Path(excel_path) if excel_path else None
        self.sheet_name = sheet_name
        self.target_label = target_label
        self.skip_label = skip_label
        self.file_extension = file_extension
        
    def rename_files(self):
        """ファイル名変更処理を実行"""
        if not self.excel_path:
            print("エラー: エクセルファイルパスが設定されていません")
            return False
            
        if not self.excel_path.exists():
            print(f"エラー: エクセルファイルが見つかりません: {self.excel_path}")
            return False

        print(f"エクセルファイルを読み込んでいます: {self.excel_path}")
        
        import datetime
        start_time = datetime.datetime.now()
        print(f"処理開始時刻: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        try:
            # エクセルブックを開く
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.sheet_name] if self.sheet_name else wb.active
            
            print(f"シート '{ws.title}' を処理中...")

            success_count = 0
            error_count = 0
            skip_count = 0

            # 2行目から最終行までループ（1行目はヘッダーと想定）
            for row in range(2, ws.max_row + 1):
                result = self._process_row(ws, row)
                
                if result == "成功":
                    success_count += 1
                elif result in ["ファイル無し", "失敗", "処理対象外"]:
                    if result == "処理対象外":
                        skip_count += 1
                    else:
                        error_count += 1

            # ファイルを保存
            wb.save(self.excel_path)
            
            end_time = datetime.datetime.now()
            duration = end_time - start_time
            
            print(f"\n処理完了:")
            print(f"  開始時刻: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"  終了時刻: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"  処理時間: {duration}")
            print(f"  ---")
            print(f"  成功: {success_count}件")
            print(f"  失敗: {error_count}件")
            print(f"  対象外: {skip_count}件")
            print(f"  合計: {success_count + error_count + skip_count}件")
            print(f"\nエクセルファイルを保存しました: {self.excel_path}")
            
            return True

        except Exception as e:
            print(f"致命的なエラーが発生しました: {str(e)}")
            return False
    
    def _process_row(self, ws, row):
        """1行の処理を実行"""
        # 各列の値を取得
        folder_path_cell = ws.cell(row=row, column=1).value
        before_name_cell = ws.cell(row=row, column=2).value
        after_name_cell  = ws.cell(row=row, column=3).value
        target_flag_cell = ws.cell(row=row, column=4).value
        
        result_cell = ws.cell(row=row, column=5)

        # 変更対象かどうかチェック
        if str(target_flag_cell).strip() != self.target_label:
            if str(target_flag_cell).strip() == self.skip_label:
                result_cell.value = "処理対象外"
            return "処理対象外"

        # 必要なデータが揃っているかチェック
        if not folder_path_cell or not before_name_cell or not after_name_cell:
            result_cell.value = "失敗: データ不足"
            print(f"[{row}行目] 失敗: A, B, C列のいずれかが空です")
            return "失敗"

        # フォルダパスとファイル名の構築
        folder_path = Path(str(folder_path_cell).strip())
        # ファイル名については、末尾のスペースが意図的な場合があるため strip() を外す
        # ただし、改行コードなどの不可視文字のみを除去する
        before_name = str(before_name_cell).replace('\r', '').replace('\n', '')
        after_name  = str(after_name_cell).replace('\r', '').replace('\n', '')
        
        # 拡張子の補完
        if not before_name.lower().endswith(self.file_extension.lower()):
            before_name += self.file_extension
        if not after_name.lower().endswith(self.file_extension.lower()):
            after_name += self.file_extension

        old_file_path = folder_path / before_name
        new_file_path = folder_path / after_name

        try:
            if not old_file_path.exists():
                result_cell.value = "ファイル無し"
                print(f"[{row}行目] ファイル無し: {old_file_path.name}")
                return "ファイル無し"
            elif new_file_path.exists():
                result_cell.value = "失敗: 変更後の名称が既に存在"
                print(f"[{row}行目] 重複エラー: {new_file_path.name}")
                return "失敗"
            else:
                # 実際の変更処理
                os.rename(old_file_path, new_file_path)
                result_cell.value = "成功"
                print(f"[{row}行目] 成功: {before_name} -> {after_name}")
                return "成功"
        
        except Exception as e:
            result_cell.value = f"失敗: {str(e)}"
            print(f"[{row}行目] エラー: {str(e)}")
            return "失敗"

def main():
    """メイン実行関数"""
    # フォルダパスとファイル名を結合
    excel_full_path = Path(EXCEL_DIR) / EXCEL_NAME
    
    # 冒ントの変数を使用してインスタンス化
    renamer = ExcelFileRenamer(
        excel_path=excel_full_path,
        sheet_name=SHEET_NAME,
        target_label=TARGET_LABEL,
        skip_label=SKIP_LABEL,
        file_extension=FILE_EXTENSION
    )
    
    # 引数があれば上書き（必要なければこの部分は削除可能）
    if len(sys.argv) > 1:
        renamer.excel_path = Path(sys.argv[1])
    
    renamer.rename_files()

if __name__ == "__main__":
    main()
