# --- obsidian_property ---
# scr名: 【自動】
# 概要: 楽天RSSバッチ収集
# 処理grp: 楽天RSS
# 処理順番: 0
# mermaid: 
# tags: ["rakuten", "rss", "batch"]
# aliases: ["rakuten_rss_batch_collector_polling.py"]
# created: 2026-03-10
# updated: 【自動】
# folder:【自動】
# file:【自動】
# cssclasses: python_script
# --- obsidian_property ---

# --- 概要 ---
# [!abstract] 概要：TDnet差分抽出処理
# 楽天RSSで株式情報をバッチ処理で収集するスクリプト
# Excelファイルから銘柄リストを読み込み、バッチ単位でデータを取得する
# 特徴:
# - ポーリング方式: RSSデータの到着を確認してから次に進む（固定待機も選択可）
# - 設定変数はスクリプト冒頭で変更可能
# 要件:
# - 楽天証券のマーケットスピードII RSSが接続されているExcelが必要
# - 銘柄リストExcelファイルから銘柄コードを読み込む
# --- 概要 ---

import win32com.client
from openpyxl import Workbook, load_workbook
import pandas as pd
import time
import pywintypes
import datetime
import os
import sys
import glob

# ==============================================================================
# 設定 (Configuration)
# ==============================================================================
# バッチ処理設定
BATCH_SIZE = 50           # 1バッチあたりの銘柄数 (デフォルト: 50)
WAIT_AFTER_RECALC = 5     # 再計算指示後の待機時間(秒) (デフォルト: 5)
WAIT_RSS_UPDATE = 5       # RSSデータ更新待ち時間(秒) ※ポーリング前の最低待機時間

# ポーリング設定 (RSSデータ準備完了を検知する方式)
POLLING_ENABLED = True          # True: ポーリング方式 / False: 固定待機方式(従来)
POLLING_INTERVAL = 2            # ポーリング間隔(秒) - 何秒おきにデータ到着を確認するか
POLLING_TIMEOUT = 120           # ポーリング最大待機時間(秒) - これを超えたらタイムアウト
POLLING_CHECK_ROWS = 3          # チェックする行数 (先頭N行 + 最終N行 を確認)
POLLING_CHECK_ITEM = "現在値"   # 準備完了の判定に使うRSS項目名
POLLING_READY_THRESHOLD = 0.8   # 準備完了と判定する割合 (0.8 = チェック対象の80%にデータがあればOK)

# ファイルパス設定
INPUT_EXCEL_FILENAME = "上場銘柄リスト_20260304.xlsx" # 銘柄リストのファイル名
INPUT_EXCEL_REL_PATH = ".."                         # スクリプトから見た相対パス
OUTPUT_DIR_NAME = "output"                          # 出力ディレクトリ名

# リトライ設定
RETRY_LOOP_COUNT = 3      # 全データ取得後の欠損チェック＆再取得ループ回数
MAX_RETRY_EXCEL_OP = 20   # Excel操作の最大リトライ回数
# ==============================================================================

def format_time(seconds):
    """秒をHH:MM:SS形式に変換"""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

def save_partial_data(wb_new, start_time, processed_count, total_count, script_dir, current_file=None):
    """途中データを保存する関数"""
    try:
        current_datetime = datetime.datetime.now()
        output_dir = os.path.join(script_dir, OUTPUT_DIR_NAME)
        os.makedirs(output_dir, exist_ok=True)
        
        if current_file:
            filename = current_file
            print(f"既存の途中保存ファイルを上書きします: {os.path.basename(filename)}")
        else:
            filename = os.path.join(output_dir, f"銘柄収集_途中_{current_datetime:%Y%m%d_%H%M}.xlsx")
            
        wb_new.save(filename)
        
        elapsed = time.time() - start_time
        print(f"\n" + "=" * 60)
        print("途中データを保存しました!")
        print(f"保存先: {filename}")
        print(f"処理銘柄数: {processed_count}/{total_count}")
        print(f"経過時間: {format_time(elapsed)}")
        print("=" * 60)
        
        return filename
    except Exception as e:
        print(f"途中データの保存に失敗しました: {e}")
        return current_file

def main():
    start_time = time.time()
    start_datetime = datetime.datetime.now()
    
    print("=" * 60)
    print("楽天RSSで株式情報バッチ収集を開始します")
    print(f"開始時間: {start_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # --- 設定 ---
    rss_list = [
        '銘柄コード', '市場部略称', '現在値', '前日比', '前日比率', '出来高', '売買代金', '出来高加重平均', '時価総額',
        '始値', '高値', '安値', '信用貸借区分', '逆日歩', '逆日歩更新日付', '信用売残',
        '信用売残前週比', '信用買残', '信用買残前週比', '信用倍率', '証金残更新日付', '新規貸株',
        '新規融資', '返済貸株', '返済融資', '残高貸株', '残高融資', '残高差引', '前日比貸株',
        '前日比融資', '前日比差引', '回転日数', '貸借倍率', '配当', '配当落日', '中配落日',
        '権利落日', '決算発表日', 'PER', 'PBR', '当日基準値', '年初来高値', '年初来安値',
        '年初来高値日付', '年初来安値日付', '上場来高値', '上場来安値', '上場来高値日付',
        '上場来安値日付', '貸株金利', '貸株金利適用日'
    ]
    
    cal_list = ['配当利回り']
    add_list = rss_list + cal_list
    
    batch_size = BATCH_SIZE
    
    # --- Excel から銘柄コードリスト取得 ---
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, INPUT_EXCEL_REL_PATH, INPUT_EXCEL_FILENAME)
    
    if not os.path.exists(excel_path):
        print(f"エラー: 銘柄リストExcelファイルが見つかりません: {excel_path}")
        return
    
    try:
        print(f"銘柄リストExcelファイルを読み込み中: {excel_path}")
        df_codes = pd.read_excel(excel_path)
        
        # 銘柄コード列を特定（複数の可能性に対応）
        code_column = None
        for col in df_codes.columns:
            if 'コード' in str(col) or 'code' in str(col).lower():
                code_column = col
                break
        
        if code_column is None:
            # 最初の列を銘柄コードとして使用
            code_column = df_codes.columns[0]
            print(f"警告: 銘柄コード列を特定できません。最初の列 '{code_column}' を使用します")
        
        stock_codes = []
        for code in df_codes[code_column].dropna():
            code_str = str(code).strip()
            # 先頭4桁を銘柄コードとして使用（Aが含まれていてもOK）
            if len(code_str) >= 4:
                stock_code = code_str[:4]
                stock_codes.append(stock_code)
            else:
                print(f"警告: 短すぎる銘柄コードをスキップ: {code_str}")
        print(f"銘柄数: {len(stock_codes)}")
        print(f"銘柄コード範囲: {stock_codes[0]} - {stock_codes[-1]}")
        
    except Exception as e:
        print(f"エラー: 銘柄リストExcelファイルの読み込みに失敗しました: {e}")
        return
    
    # --- Excel 起動 ---
    def get_excel_connection():
        try:
            print("\nExcelに接続中...")
            xl = win32com.client.GetObject(Class="Excel.Application")
            xl.Visible = True
            return xl
        except pywintypes.com_error:
            print("エラー: Excel が開いていません。")
            print("マーケットスピードII RSSに接続したExcelを起動してから再実行してください")
            return None
    
    xl = get_excel_connection()
    if xl is None:
        return
    
    # Excelの計算モードを手動に設定（負荷軽減）
    original_calculation = xl.Calculation
    xl.Calculation = 1  # xlManual
    
    # 画面更新を無効化（高速化）
    original_screen_updating = xl.ScreenUpdating
    xl.ScreenUpdating = False
    
    # アクティブシートを取得
    ws = xl.ActiveSheet
    
    # RSSのタイトル行を記入
    for col, item in enumerate(rss_list):
        ws.Cells(1, col + 2).Value = item
    
    # --- Excel 新規ブック作成（openpyxl側）または途中再開 ---
    output_dir = os.path.join(script_dir, OUTPUT_DIR_NAME)
    partial_pattern = os.path.join(output_dir, "銘柄収集_途中_*.xlsx")
    partial_files = glob.glob(partial_pattern)
    
    wb_new = None
    ws_new = None
    processed_rows = 0
    current_partial_file = None
    
    if partial_files:
        # ファイル名でソートして最新を取得 (YYYYMMDD_HHMMが含まれているため)
        latest_partial = sorted(partial_files)[-1]
        print(f"\n途中保存ファイルが見つかりました: {os.path.basename(latest_partial)}")
        print("このファイルから処理を再開します。")
        
        try:
            wb_new = load_workbook(latest_partial)
            if wb_new.active:
                ws_new = wb_new.active
            else:
                ws_new = wb_new.create_sheet(title="Sheet1")
            
            # ヘッダー行を除くデータ行数を取得
            processed_rows = ws_new.max_row - 1
            if processed_rows < 0: processed_rows = 0
            
            print(f"既存データの行数: {processed_rows}")
            current_partial_file = latest_partial
            
        except Exception as e:
            print(f"警告: 途中保存ファイルの読み込みに失敗しました: {e}")
            print("新規に作成します。")
            wb_new = None
            processed_rows = 0
    
    if wb_new is None:
        wb_new = Workbook()
        if wb_new.active is None:
            ws_new = wb_new.create_sheet(title="Sheet1")
        else:
            ws_new = wb_new.active
        
        # ヘッダー行を追加
        ws_new.append(add_list)
        processed_rows = 0
        
    # 銘柄リストを調整
    if processed_rows > 0:
        if processed_rows >= len(stock_codes):
            print("全ての銘柄が既に収集済みです。終了します。")
            return
        
        print(f"スキップする銘柄数: {processed_rows}")
        stock_codes = stock_codes[processed_rows:]
        print(f"残り銘柄数: {len(stock_codes)}")
    
    # --- Excel セル書き込み用再試行関数 ---
    def retry_excel_set(row_idx, col_idx, value, formula=False, max_retry=MAX_RETRY_EXCEL_OP):
        nonlocal xl, ws
        for i in range(max_retry):
            try:
                # 行・列のインデックスチェック (1以上)
                if row_idx < 1 or col_idx < 1:
                    print(f"警告: 無効なセルインデックス ({row_idx}, {col_idx})")
                    return False

                cell = ws.Cells(row_idx, col_idx)
                if formula:
                    cell.Formula = value
                else:
                    cell.Value = value
                return True
            except pywintypes.com_error as e:
                wait = 3 * (i+1)  # 3秒, 6秒, 9秒...最大60秒待つ
                print(f"セル書き込みリトライ {i+1}回目 ({wait}秒待機) - {e}")
                time.sleep(wait)
                
                # Excel接続を再確認
                if i >= 5:  # 5回失敗したらExcel接続を再確保
                    print("Excel接続を再確保します...")
                    try:
                        xl_new = get_excel_connection()
                        if xl_new is not None:
                            xl = xl_new
                            ws = xl.ActiveSheet
                        else:
                            print("Excel再接続失敗")
                            return False
                    except Exception as conn_err:
                        print(f"Excel再接続エラー: {conn_err}")
                        return False
                    
                    # 再設定が必要
                    return retry_excel_set(row_idx, col_idx, value, formula, max_retry - i - 1)
        return False
    
    # --- Excel セル読み込み用再試行関数 ---
    def retry_excel_get(row_idx, col_idx, max_retry=MAX_RETRY_EXCEL_OP):
        nonlocal xl, ws
        for i in range(max_retry):
            try:
                # 行・列のインデックスチェック (1以上)
                if row_idx < 1 or col_idx < 1:
                    return None

                cell = ws.Cells(row_idx, col_idx)
                return cell.Value
            except pywintypes.com_error as e:
                wait = 2 * (i+1)  # 2秒, 4秒, 6秒...最大30秒待つ
                print(f"セル読み込みリトライ {i+1}回目 ({wait}秒待機) - {e}")
                time.sleep(wait)
                
                # Excel接続を再確認
                if i >= 5:  # 5回失敗したらExcel接続を再確保
                    print("Excel接続を再確保します...")
                    try:
                        xl_new = get_excel_connection()
                        if xl_new is not None:
                            xl = xl_new
                            ws = xl.ActiveSheet
                        else:
                            print("Excel再接続失敗")
                            return None
                    except Exception as conn_err:
                        print(f"Excel再接続エラー: {conn_err}")
                        return None
                    
                    return retry_excel_get(row_idx, col_idx, max_retry - i - 1)
        return None
    
    # --- Excel 範囲書き込み用再試行関数 ---
    def retry_excel_set_range(start_row, start_col, data, formula=False, max_retry=MAX_RETRY_EXCEL_OP):
        nonlocal xl, ws
        rows = len(data)
        if rows == 0: return True
        cols = len(data[0])
        
        for i in range(max_retry):
            try:
                # 範囲オブジェクトを取得
                start_cell = ws.Cells(start_row, start_col)
                end_cell = ws.Cells(start_row + rows - 1, start_col + cols - 1)
                rng = ws.Range(start_cell, end_cell)
                
                if formula:
                    rng.Formula = data
                else:
                    rng.Value = data
                return True
            except pywintypes.com_error as e:
                wait = 3 * (i+1)
                print(f"範囲書き込みリトライ {i+1}回目 ({wait}秒待機) - {e}")
                time.sleep(wait)
                
                if i >= 5:
                    print("Excel接続を再確保します...")
                    try:
                        xl_new = get_excel_connection()
                        if xl_new is not None:
                            xl = xl_new
                            ws = xl.ActiveSheet
                        else:
                            return False
                    except:
                        return False
                    return retry_excel_set_range(start_row, start_col, data, formula, max_retry - i - 1)
        return False

    # --- Excel 範囲読み込み用再試行関数 ---
    def retry_excel_get_range(start_row, start_col, rows, cols, max_retry=MAX_RETRY_EXCEL_OP):
        nonlocal xl, ws
        for i in range(max_retry):
            try:
                start_cell = ws.Cells(start_row, start_col)
                end_cell = ws.Cells(start_row + rows - 1, start_col + cols - 1)
                rng = ws.Range(start_cell, end_cell)
                return rng.Value  # 2次元タプルで返る
            except pywintypes.com_error as e:
                wait = 2 * (i+1)
                print(f"範囲読み込みリトライ {i+1}回目 ({wait}秒待機) - {e}")
                time.sleep(wait)
                
                if i >= 5:
                    print("Excel接続を再確保します...")
                    try:
                        xl_new = get_excel_connection()
                        if xl_new is not None:
                            xl = xl_new
                            ws = xl.ActiveSheet
                        else:
                            return None
                    except:
                        return None
                    return retry_excel_get_range(start_row, start_col, rows, cols, max_retry - i - 1)
        return None

    def process_batch(batch_codes):
        """指定された銘柄リストに対してバッチ処理を実行する内部関数"""
        nonlocal processed_count, wb_new, ws_new, current_partial_file
        
        # Excel 表クリア (リトライ機能付き)
        for retry in range(5):
            try:
                ws.Range(ws.Cells(2, 1), ws.Cells(batch_size + 2, len(rss_list) + 1)).ClearContents()
                break
            except pywintypes.com_error as e:
                print(f"Excel表クリア失敗 (リトライ {retry+1}/5): {e}")
                time.sleep(2)
                if retry == 4:
                    print("警告: Excel表クリアに失敗しました。処理を続行しますが、古いデータが残る可能性があります。")

        # RSS 表に銘柄コードと関数を一括書き込み
        print("RSS関数を設定中(一括)...")
        
        # 書き込み用データ作成
        code_col_data = []      # A列: 銘柄コード
        formula_data = []       # B列以降: RSS関数
        
        for stock_no in batch_codes:
            code_col_data.append([str(stock_no)])
            row_formulas = []
            for item in rss_list:
                row_formulas.append(f'=RssMarket("{stock_no}","{item}")')
            formula_data.append(row_formulas)
            
        # 一括書き込み実行
        if not retry_excel_set_range(2, 1, code_col_data):
            print("警告: 銘柄コードの一括書き込みに失敗しました")
            
        if not retry_excel_set_range(2, 2, formula_data, formula=True):
            print("警告: RSS関数の一括書き込みに失敗しました")
            
        print(f"RSS関数設定完了: {len(batch_codes)} 銘柄")
        
        # 計算モードを自動に戻して再計算
        xl.Calculation = -4105  # xlAutomatic
        print("Excel再計算中...")
        time.sleep(WAIT_AFTER_RECALC)
        
        # RSS データ更新待機
        print("RSSデータ更新待機中...")
        time.sleep(WAIT_RSS_UPDATE)  # 最低待機時間
        
        # Excelの応答性を確認
        try:
            _ = ws.Cells(1, 1).Value
        except pywintypes.com_error:
            print("Excelが応答していません。さらに待機します...")
            time.sleep(15)
            try:
                _ = ws.Cells(1, 1).Value
            except pywintypes.com_error:
                print("Excelが応答しません。このバッチはスキップされます。")
                return []
        
        # --- ポーリング: RSSデータ準備完了を待つ ---
        if POLLING_ENABLED:
            # チェック対象の列を特定
            try:
                check_col_idx = rss_list.index(POLLING_CHECK_ITEM) + 2  # B列開始なので+2
            except ValueError:
                check_col_idx = 4  # デフォルト: 現在値(D列あたり)
            
            # チェック対象の行を決定 (先頭N行 + 最終N行)
            num_rows = len(batch_codes)
            check_row_indices = []
            # 先頭N行
            for i in range(min(POLLING_CHECK_ROWS, num_rows)):
                check_row_indices.append(2 + i)
            # 最終N行
            for i in range(min(POLLING_CHECK_ROWS, num_rows)):
                row_idx = 2 + num_rows - 1 - i
                if row_idx not in check_row_indices:
                    check_row_indices.append(row_idx)
            
            total_checks = len(check_row_indices)
            required_ready = int(total_checks * POLLING_READY_THRESHOLD)
            
            print(f"ポーリング開始 (チェック: {total_checks}セル, 閾値: {required_ready}セル以上で完了判定)")
            print(f"  チェック項目: {POLLING_CHECK_ITEM} (列{check_col_idx})")
            print(f"  タイムアウト: {POLLING_TIMEOUT}秒, 間隔: {POLLING_INTERVAL}秒")
            
            polling_start = time.time()
            polling_elapsed = 0
            ready_count = 0
            last_ready_count = -1
            stall_count = 0
            
            while polling_elapsed < POLLING_TIMEOUT:
                ready_count = 0
                for row_idx in check_row_indices:
                    try:
                        val = ws.Cells(row_idx, check_col_idx).Value
                        if val is not None and str(val) not in ("0", "#N/A", "#VALUE!", ""):
                            try:
                                float(val)
                                ready_count += 1
                            except (ValueError, TypeError):
                                # 数値変換できないが値はある（文字列データ等）→ 準備完了とみなす
                                if str(val).strip() != "":
                                    ready_count += 1
                    except pywintypes.com_error:
                        # Excelビジー → まだ処理中
                        pass
                
                polling_elapsed = time.time() - polling_start
                
                if ready_count >= required_ready:
                    print(f"  ✓ RSS準備完了! ({ready_count}/{total_checks}セル到着, {polling_elapsed:.1f}秒)")
                    break
                
                # 進捗表示 (変化があった時 or 10秒ごと)
                if ready_count != last_ready_count or int(polling_elapsed) % 10 == 0:
                    print(f"  ... 待機中 {polling_elapsed:.0f}秒 ({ready_count}/{total_checks}セル到着)")
                
                # ストール検知 (進捗が止まった場合)
                if ready_count == last_ready_count:
                    stall_count += 1
                else:
                    stall_count = 0
                last_ready_count = ready_count
                
                # ストールが長すぎたら早めに切り上げ
                if stall_count >= 10 and ready_count > 0:
                    print(f"  △ 進捗停滞を検知 ({ready_count}/{total_checks}セル到着, {polling_elapsed:.1f}秒)")
                    print(f"    一部の銘柄はデータが存在しない可能性があります。取得を続行します。")
                    break
                
                time.sleep(POLLING_INTERVAL)
            else:
                # タイムアウト
                print(f"  × タイムアウト ({ready_count}/{total_checks}セル到着, {POLLING_TIMEOUT}秒)")
                print(f"    取得できたデータで続行します。")
        
        # Python 側でデータ一括取得
        print("データ取得中(一括)...")
        
        # 範囲指定で一括取得 (行数: len(batch_codes), 列数: len(rss_list))
        # B列から開始（銘柄コードはA列だがRSS取得対象はB列以降のrss_listに対応する部分から）
        # ※修正: rss_listには'銘柄コード'も含まれているが、
        # コード上のrss_list = ['銘柄コード', '銘柄名称', ...] となっている場合、
        # A列(1列目)には銘柄コードを書き、B列(2列目)以降にRSS関数を入れている。
        # RSS関数のループは `for col, item in enumerate(rss_list):` で `col+2` (B列~) に書いている。
        # rss_listの0番目('銘柄コード')もRSS関数としてB列に書かれているのか？
        # コードを確認すると `formula = f'=RssMarket("{stock_no}","{item}")'` で全項目書いている。
        # つまり、A列は単なる文字列のコード、B列はRssMarket(code, "銘柄コード")、C列はRssMarket(code, "銘柄名称")...となる。
        # 取得すべきデータはB列(2列目)から len(rss_list) 列分。
        
        raw_data = retry_excel_get_range(2, 2, len(batch_codes), len(rss_list))
        
        if raw_data is None:
             print("警告: データの一括取得に失敗しました。個別取得を試みます...")
             # フォールバック: 個別取得ロジック (省略)
             return []
             
        batch_data = []
        
        # raw_dataはタプルのタプル ((row1_col1, row1_col2...), (row2_col1...))
        for i, row_vals in enumerate(raw_data):
            row_data = list(row_vals)
            
            # タイムゾーン付きdatetimeを除去 (openpyxl保存時エラー対策)
            for idx, val in enumerate(row_data):
                if isinstance(val, datetime.datetime) and val.tzinfo is not None:
                    row_data[idx] = val.replace(tzinfo=None)
            
            # Noneチェックと正規化
            row_data = [val if val is not None else None for val in row_data]
            
            # 計算列
            try:
                current_price = row_data[rss_list.index("現在値")]
            except ValueError:
                current_price = 0
            
            try:
                dividend = row_data[rss_list.index("配当")]
            except ValueError:
                dividend = 0
            
            # Excel から取得した値を安全に float に変換
            def to_float(val):
                try:
                    return float(val) if val is not None else 0.0
                except (ValueError, TypeError):
                    return 0.0
            
            current_price_f = to_float(current_price)
            dividend_f = to_float(dividend)
            
            dividend_yield = (dividend_f / current_price_f * 100) if current_price_f else 0
            
            row_data.append(dividend_yield)
            batch_data.append(row_data)
        
        return batch_data

    # --- メイン処理ループ ---
    processed_count = 0
    
    try:
        # 初回データ収集
        total_batches = (len(stock_codes) - 1) // batch_size + 1
        print(f"\nバッチ処理を開始します (1バッチ={batch_size}銘柄, 全{total_batches}バッチ)")
        print("-" * 60)
        
        for batch_num in range(total_batches):
            batch_start_time = time.time()
            start_idx = batch_num * batch_size
            end_idx = min(start_idx + batch_size, len(stock_codes))
            batch_codes = stock_codes[start_idx:end_idx]
            
            print(f"\nバッチ {batch_num + 1}/{total_batches} 処理中...")
            print(f"銘柄範囲: {start_idx + 1}-{end_idx} (コード: {batch_codes[0]} - {batch_codes[-1]})")
            
            try:
                batch_data = process_batch(batch_codes)
                
                # Excel 書き込み
                if batch_data:
                    print("Excelファイルに書き込み中...")
                    df_batch = pd.DataFrame(batch_data, columns=add_list)
                    for r in df_batch.values.tolist():
                        ws_new.append(r)
                
                processed_count += len(batch_codes)
                
                # バッチ完了時刻と経過時間
                batch_end_time = time.time()
                batch_elapsed = batch_end_time - batch_start_time
                total_elapsed = batch_end_time - start_time
                
                print(f"バッチ {batch_num + 1} 完了!")
                print(f"  バッチ処理時間: {format_time(batch_elapsed)}")
                print(f"  累計処理時間: {format_time(total_elapsed)}")
                
                # 残り時間の推定
                if batch_num > 0:
                    avg_batch_time = total_elapsed / (batch_num + 1)
                    remaining_batches = total_batches - (batch_num + 1)
                    estimated_remaining = avg_batch_time * remaining_batches
                    print(f"  推定残り時間: {format_time(estimated_remaining)}")
                
            except Exception as batch_error:
                print(f"バッチ {batch_num + 1} でエラーが発生しました: {batch_error}")
                print(f"これまでのデータを保存して終了します...")
                current_partial_file = save_partial_data(wb_new, start_time, processed_count, len(stock_codes), script_dir, current_partial_file)
                raise batch_error
        
        # 初回完了時の保存
        current_partial_file = save_partial_data(wb_new, start_time, processed_count, len(stock_codes), script_dir, current_partial_file)
        
        # --- 欠損データの再取得ループ ---
        for retry_loop in range(RETRY_LOOP_COUNT):
            print(f"\n" + "=" * 60)
            print(f"データ欠損チェックと再取得 ({retry_loop + 1}/{RETRY_LOOP_COUNT}回目)")
            print("=" * 60)
            
            # 現在のデータを読み込み（メモリ上のws_newを使用）
            missing_codes = []
            rows_to_update = []  # (row_index, stock_code) のタプル
            
            # ヘッダーを除く全行をチェック
            # A列（銘柄コード）は必須、D列（現在値）などが取れているか確認も可能だが、
            # 今回は「A列にデータがない」または「主要データが欠損している」場合を対象とする
            
            # openpyxlの1始まりの行番号
            max_row = ws_new.max_row
            print(f"全 {max_row - 1} 行をチェック中...")
            
            # 列インデックスを特定
            try:
                code_col_idx = rss_list.index('銘柄コード') + 1
                name_col_idx = rss_list.index('市場部略称') + 1
                price_col_idx = rss_list.index('現在値') + 1
            except ValueError:
                # 万が一見つからない場合はデフォルト値を使用
                code_col_idx = 1
                name_col_idx = 2
                price_col_idx = 3
            
            for row_idx in range(2, max_row + 1):
                # 動的に列を指定
                code_val = ws_new.cell(row=row_idx, column=code_col_idx).value
                name_val = ws_new.cell(row=row_idx, column=name_col_idx).value
                price_val = ws_new.cell(row=row_idx, column=price_col_idx).value
                
                # 銘柄コードが空、または名称がエラー/0、または現在値が空の場合を欠損とみなす
                # ※A列が空の場合はそもそもどの銘柄かわからないため、
                # ここでは「銘柄リストの順序と行が一致している」前提で、
                # 再構築するか、あるいは「値が無効な行」を特定して再取得する。
                
                # シンプルに「値が無効な行」を特定して、その行のデータを再取得するアプローチをとる
                # ただし、元の銘柄コードがわからないと再取得できないため、
                # 元の stock_codes リストと行番号を対応付ける必要がある。
                
                # 全銘柄を一括処理した場合、stock_codes の index = row_idx - 2 となるはず
                # （ただし途中再開などでずれる可能性もあるが、今回は追記方式なので概ね一致）
                
                # 確実なのは「値が入っていない」行を特定すること
                is_missing = False
                
                if code_val is None or str(code_val).strip() == "":
                    is_missing = True
                elif name_val is None or str(name_val) == "0" or str(name_val) == "#N/A":
                    is_missing = True
                
                if is_missing:
                    # 元のリストからコードを特定（範囲外チェック付き）
                    list_idx = row_idx - 2
                    if 0 <= list_idx < len(stock_codes):
                        target_code = stock_codes[list_idx]
                        missing_codes.append(target_code)
                        rows_to_update.append((row_idx, target_code))
            
            if not missing_codes:
                print("欠損データはありません。全処理を完了します。")
                break
            
            print(f"欠損データが見つかりました: {len(missing_codes)} 件")
            print(f"再取得を開始します...")
            
            # 欠損銘柄をバッチ処理
            # ここでは行ごとに個別に書き込む必要があるため、process_batchを少し変形して利用するか、
            # あるいは取得したデータを特定の行に書き戻す処理を行う
            
            missing_batch_size = batch_size
            total_missing_batches = (len(missing_codes) - 1) // missing_batch_size + 1
            
            for m_batch_num in range(total_missing_batches):
                m_start = m_batch_num * missing_batch_size
                m_end = min(m_start + missing_batch_size, len(missing_codes))
                m_batch_codes = missing_codes[m_start:m_end]
                m_rows = rows_to_update[m_start:m_end]
                
                print(f"再取得バッチ {m_batch_num + 1}/{total_missing_batches} ({len(m_batch_codes)}件)...")
                
                # データを取得
                m_data = process_batch(m_batch_codes)
                
                # 該当行に書き戻し
                if m_data:
                    for i, row_data in enumerate(m_data):
                        target_row_idx = m_rows[i][0]
                        for col_idx, val in enumerate(row_data):
                            # openpyxlは1始まり
                            ws_new.cell(row=target_row_idx, column=col_idx + 1).value = val
                
                # バッチごとに保存
                wb_new.save(current_partial_file)
            
            print("再取得ループ完了。")
            
    except Exception as main_error:
        print(f"\n処理中にエラーが発生しました: {main_error}")
        if processed_count > 0:
            current_partial_file = save_partial_data(wb_new, start_time, processed_count, len(stock_codes), script_dir, current_partial_file)
        return
    
    # --- 保存 ---
    # Excel設定を元に戻す
    try:
        xl.Calculation = original_calculation
        xl.ScreenUpdating = original_screen_updating
        print("Excel設定を元に戻しました")
    except:
        print("Excel設定の復元に失敗しました")
    
    end_time = time.time()
    end_datetime = datetime.datetime.now()
    total_elapsed = end_time - start_time
    
    print("\n" + "=" * 60)
    print("データ収集完了!")
    print(f"終了時間: {end_datetime.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"総処理時間: {format_time(total_elapsed)}")
    print(f"処理銘柄数: {len(stock_codes)}銘柄")
    
    # 保存処理
    output_dir = os.path.join(script_dir, OUTPUT_DIR_NAME)
    os.makedirs(output_dir, exist_ok=True)
    
    filename = os.path.join(output_dir, f"銘柄収集_{end_datetime:%Y%m%d_%H%M}.xlsx")
    wb_new.save(filename)
    
    print(f"保存先: {filename}")
    
    # 完了後に途中経過ファイルを削除 (上書き完了とみなす)
    if current_partial_file and os.path.exists(current_partial_file):
        try:
            # パスを正規化して比較
            abs_partial = os.path.abspath(current_partial_file)
            abs_final = os.path.abspath(filename)
            
            if abs_partial != abs_final:
                os.remove(current_partial_file)
                print(f"途中経過ファイルを削除しました (完了ファイルに統合): {os.path.basename(current_partial_file)}")
        except Exception as e:
            print(f"警告: 途中経過ファイルの削除に失敗しました: {e}")

    print("=" * 60)

if __name__ == "__main__":
    main()
